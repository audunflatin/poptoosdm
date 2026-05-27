from fastapi import FastAPI, Request, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, StreamingResponse, Response
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from starlette.types import Scope

from pathlib import Path
import csv
import decimal
import io
import json
import math

import base64
import threading
import uuid
import requests
import ijson

from backend.auth_db import SessionLocal, User, LoginLog, PasswordResetToken, EventLog, AccessRequest, init_db
from backend.auth_utils import verify_password, generate_password, hash_password
from backend.core.settings import SESSION_SECRET, ALLOWED_ORIGINS
from backend.email_utils import send_welcome_email, send_reset_email, send_reset_link_email, send_contact_email, send_access_request_email

import logging
logger = logging.getLogger(__name__)

from collections import defaultdict
import time

from zoneinfo import ZoneInfo
from datetime import datetime, timezone, timedelta

# ---------------------------------------------------------------------
# App
# ---------------------------------------------------------------------

app = FastAPI(title="OSDMTools")


class _NoHTMLStaticFiles(StaticFiles):
    """Block direct access to .html files — they must go through Python routes."""
    async def get_response(self, path: str, scope: Scope) -> Response:
        if path.endswith(".html"):
            raise HTTPException(status_code=404)
        return await super().get_response(path, scope)


app.mount("/static", _NoHTMLStaticFiles(directory="frontend"), name="static")

app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET, same_site="strict")
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["GET", "POST"],
    allow_headers=["Content-Type"],
)

# ---------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------

@app.on_event("startup")
def startup():
    init_db()
    if SESSION_SECRET == "CHANGE_ME_BEFORE_PROD":
        logger.warning("⚠️  SESSION_SECRET er ikke satt – sesjoner er ikke sikre!")

# ---------------------------------------------------------------------
# Rate limiting (innlogging)
# ---------------------------------------------------------------------

_login_attempts: dict[str, list[float]] = defaultdict(list)
_LOGIN_WINDOW = 60   # sekunder
_LOGIN_MAX    = 10   # maks forsøk per vindu

def _get_client_ip(request: Request) -> str:
    """Hent reell klient-IP — foretrekker Cloudflare-header."""
    return (
        request.headers.get("CF-Connecting-IP")
        or (request.client.host if request.client else "unknown")
    )

def _rate_limit_check(ip: str):
    now = time.time()
    _login_attempts[ip] = [t for t in _login_attempts[ip] if now - t < _LOGIN_WINDOW]
    if len(_login_attempts[ip]) >= _LOGIN_MAX:
        raise HTTPException(status_code=429, detail="For mange innloggingsforsøk. Prøv igjen om litt.")
    _login_attempts[ip].append(now)

def _rate_limit_reset(ip: str):
    _login_attempts.pop(ip, None)

_access_requests: dict[str, list[float]] = defaultdict(list)
_ACCESS_WINDOW = 86400  # 24 timer
_ACCESS_MAX    = 3      # maks forespørsler per dag per IP

def _check_access_rate_limit(ip: str):
    now = time.time()
    _access_requests[ip] = [t for t in _access_requests[ip] if now - t < _ACCESS_WINDOW]
    if len(_access_requests[ip]) >= _ACCESS_MAX:
        raise HTTPException(status_code=429, detail="rate_limited")
    _access_requests[ip].append(now)

# ---------------------------------------------------------------------
# State
# ---------------------------------------------------------------------

TEN_TABLE = None
GENERATION_PROGRESS = {"status": "idle", "percent": 0}
OSDM_IN = Path("data/input/1076-OSDM-template.json")
OSDM_STORE: dict[str, dict] = {}     # user_email → {"filename": str, "content": str}
FIX_OSDM_STORE: dict[str, dict] = {} # user_email → {"filename": str, "content": bytes}
XLSX_JOBS: dict = {}        # job_id → {status, result, error, filename, percent, rows, owner, created_at}
VALIDATION_JOBS: dict = {}  # job_id → {status, percent, phase, start_time, file_size, result, error, owner}
PARSE_JOBS: dict = {}       # job_id → {status, percent, phase, start_time, file_size, result, error, owner}

_JOB_TTL = 2 * 3600  # 2 timer

def _cleanup_jobs():
    while True:
        time.sleep(600)
        cutoff = time.time() - _JOB_TTL
        for jobs in (XLSX_JOBS, VALIDATION_JOBS, PARSE_JOBS, OSDM_STORE, FIX_OSDM_STORE):
            stale = [k for k, v in list(jobs.items()) if v.get("created_at", 0) < cutoff]
            for k in stale:
                jobs.pop(k, None)

threading.Thread(target=_cleanup_jobs, daemon=True).start()


def _safe_filename(name: str) -> str:
    """Fjern tegn som kan ødelegge Content-Disposition-headere."""
    return "".join(c for c in name if c not in '"\\/:*?<>|\r\n').strip() or "osdm.json"

# ---------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------

def require_login(request: Request):
    if "user_email" not in request.session:
        raise HTTPException(status_code=401, detail="Ikke innlogget")

def _check_is_admin_from_db(request: Request) -> bool:
    """Les is_admin direkte fra DB og oppdater sesjonen. Returnerer True/False."""
    user_email = request.session.get("user_email")
    if not user_email:
        return False
    with SessionLocal() as db:
        user = db.query(User).filter(User.email == user_email).first()
        is_admin = bool(user and user.is_admin)
    request.session["is_admin"] = is_admin
    return is_admin

def require_admin(request: Request):
    require_login(request)
    if not _check_is_admin_from_db(request):
        raise HTTPException(status_code=403, detail="Ikke administrator")

# ---------------------------------------------------------------------
# Event logging
# ---------------------------------------------------------------------

def log_event(
    user_email: str | None,
    event_type: str,
    status: str = "ok",
    detail: dict | None = None,
):
    try:
        db = SessionLocal()
        try:
            db.add(EventLog(
                user_email=user_email,
                event_type=event_type,
                status=status,
                detail=json.dumps(detail, ensure_ascii=False) if detail else None,
            ))
            db.commit()
        finally:
            db.close()
    except Exception:
        pass  # Logging skal aldri bryte en forespørsel

# ---------------------------------------------------------------------
# Root / GUI
# ---------------------------------------------------------------------

@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    if "user_email" not in request.session:
        return HTMLResponse(Path("frontend/landing.html").read_text(encoding="utf-8"))
    is_admin = bool(request.session.get("is_admin"))
    html = Path("frontend/index.html").read_text(encoding="utf-8")
    html = html.replace(
        "</head>",
        f"<script>window.IS_ADMIN = {str(is_admin).lower()};</script></head>"
    )
    return HTMLResponse(html)


@app.get("/personvern", response_class=HTMLResponse)
def personvern_page():
    return HTMLResponse(Path("frontend/personvern.html").read_text(encoding="utf-8"))

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if "user_email" in request.session:
        return RedirectResponse("/", status_code=302)
    return HTMLResponse(Path("frontend/login.html").read_text(encoding="utf-8"))

# ---------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------

@app.get("/health")
def health():
    return {"status": "ok"}

@app.get("/robots.txt", response_class=Response)
def robots_txt():
    content = (
        "User-agent: *\n"
        "Allow: /\n"
        "Allow: /login\n"
        "Disallow: /admin/\n"
        "Disallow: /ui/\n"
        "Disallow: /logout\n"
        "Disallow: /endre-passord\n"
        "Disallow: /price-adjust\n"
        "Disallow: /osdmtoexcel\n"
        "Disallow: /fare-discount\n"
        "Disallow: /fix-osdm\n"
        "Disallow: /kontakt\n"
        "Disallow: /request-access\n"
        "\n"
        "Sitemap: https://osdmtools.com/sitemap.xml\n"
    )
    return Response(content=content, media_type="text/plain")

@app.get("/sitemap.xml", response_class=Response)
def sitemap_xml():
    content = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        '  <url>\n'
        '    <loc>https://osdmtools.com/</loc>\n'
        '    <changefreq>monthly</changefreq>\n'
        '    <priority>1.0</priority>\n'
        '  </url>\n'
        '</urlset>\n'
    )
    return Response(content=content, media_type="application/xml")

# ---------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------

@app.post("/login")
def login(request: Request, email: str = Form(...), password: str = Form(...)):
    ip = _get_client_ip(request)
    _rate_limit_check(ip)
    email = email.lower()
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == email).first()
        if not user or not user.is_active or not verify_password(password, user.password_hash):
            db.add(LoginLog(email=email, ip_address=ip, success=False))
            db.commit()
            log_event(email, "login_failed", "error", {"ip": ip})
            raise HTTPException(status_code=401, detail="Ugyldig innlogging")
        _rate_limit_reset(ip)
        request.session["user_email"] = user.email
        request.session["is_admin"] = user.is_admin
        db.add(LoginLog(email=user.email, ip_address=ip, success=True))
        log_event(user.email, "login_success", "ok", {"ip": ip})
        if user.must_change_password:
            db.commit()
            return RedirectResponse("/change-password", status_code=302)
        if user.first_login_at is None:
            user.first_login_at = datetime.now(timezone.utc)
        db.commit()
    finally:
        db.close()
    return RedirectResponse("/price-adjust", status_code=302)

@app.get("/logout")
def logout(request: Request):
    email = request.session.get("user_email")
    request.session.clear()
    log_event(email, "logout")
    return RedirectResponse("/", status_code=302)

# ---------------------------------------------------------------------
# Bytt passord (tvinges ved første innlogging)
# ---------------------------------------------------------------------

@app.get("/change-password", response_class=HTMLResponse)
def change_password_page(request: Request):
    if "user_email" not in request.session:
        return RedirectResponse("/", status_code=302)
    return HTMLResponse(Path("frontend/change_password.html").read_text(encoding="utf-8"))

@app.post("/change-password")
def change_password(
    request: Request,
    password: str = Form(...),
    confirm: str = Form(...),
):
    if "user_email" not in request.session:
        raise HTTPException(status_code=401, detail="Ikke innlogget")
    if password != confirm:
        raise HTTPException(status_code=400, detail="Passordene er ikke like")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Passordet må være minst 8 tegn")
    email = request.session["user_email"]
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == email).first()
        if not user:
            raise HTTPException(status_code=404, detail="Bruker ikke funnet")
        user.password_hash = hash_password(password)
        user.must_change_password = False
        if user.first_login_at is None:
            user.first_login_at = datetime.now(timezone.utc)
        db.commit()
    finally:
        db.close()
    log_event(email, "password_changed")
    return {"ok": True}

# ---------------------------------------------------------------------
# Glemt passord
# ---------------------------------------------------------------------

@app.get("/forgot-password", response_class=HTMLResponse)
def forgot_password_page():
    return HTMLResponse(Path("frontend/forgot_password.html").read_text(encoding="utf-8"))

@app.post("/forgot-password")
def forgot_password(email: str = Form(...)):
    email = email.lower()
    from backend.core.settings import APP_URL as _APP_URL
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == email, User.is_active == True).first()
        if user:
            token = str(uuid.uuid4())
            expires_at = datetime.now(timezone.utc).replace(tzinfo=None) + __import__("datetime").timedelta(hours=1)
            db.add(PasswordResetToken(token=token, email=email, expires_at=expires_at))
            db.commit()
            try:
                send_reset_link_email(email, f"{_APP_URL}/reset-password/{token}")
            except Exception as exc:
                logger.error("Kunne ikke sende reset-lenke til %s: %s", email, exc)
    finally:
        db.close()
    return {"ok": True}

@app.get("/reset-password/{token}", response_class=HTMLResponse)
def reset_password_page(token: str):
    db = SessionLocal()
    try:
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        rt = db.query(PasswordResetToken).filter(
            PasswordResetToken.token == token,
            PasswordResetToken.expires_at > now,
        ).first()
        if not rt:
            return HTMLResponse("<h2>Lenken er ugyldig eller utløpt.</h2>", status_code=400)
    finally:
        db.close()
    html = Path("frontend/reset_password.html").read_text(encoding="utf-8")
    return HTMLResponse(html.replace("__TOKEN__", token))

@app.post("/reset-password/{token}")
def reset_password(token: str, password: str = Form(...), confirm: str = Form(...)):
    if password != confirm:
        raise HTTPException(status_code=400, detail="Passordene er ikke like")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Passordet må være minst 8 tegn")
    db = SessionLocal()
    try:
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        rt = db.query(PasswordResetToken).filter(
            PasswordResetToken.token == token,
            PasswordResetToken.expires_at > now,
        ).first()
        if not rt:
            raise HTTPException(status_code=400, detail="Lenken er ugyldig eller utløpt")
        user = db.query(User).filter(User.email == rt.email).first()
        if not user:
            raise HTTPException(status_code=404, detail="Bruker ikke funnet")
        user.password_hash = hash_password(password)
        user.must_change_password = False
        db.delete(rt)
        db.commit()
    finally:
        db.close()
    log_event(rt.email, "password_changed", detail={"via": "reset_link"})
    return {"ok": True}

# ---------------------------------------------------------------------
# TEN validation
# ---------------------------------------------------------------------

def validate_ten_csv(text: str):
    reader = csv.reader(io.StringIO(text), delimiter=";")
    rows = []

    for i, row in enumerate(reader, start=1):
        if len(row) != 3:
            return {"ok": False, "error": f"Linje {i}: feil kolonneantall (forventet 3, fikk {len(row)})"}

        frm, to, price = [c.strip().replace(" ", "").replace("\xa0", "") for c in row]

        try:
            frm_int = int(frm)
        except ValueError:
            return {"ok": False, "error": f"Linje {i}: ugyldig 'fra'-verdi: '{frm}' er ikke et heltall"}

        try:
            to_int = int(to)
        except ValueError:
            return {"ok": False, "error": f"Linje {i}: ugyldig 'til'-verdi: '{to}' er ikke et heltall"}

        try:
            price_int = int(price)
        except ValueError:
            return {"ok": False, "error": f"Linje {i}: ugyldig pris: '{price}' er ikke et heltall"}

        if frm_int < 0 or to_int < 0:
            return {"ok": False, "error": f"Linje {i}: negative km-verdier er ikke tillatt"}

        if frm_int >= to_int:
            return {"ok": False, "error": f"Linje {i}: 'fra' ({frm_int}) må være mindre enn 'til' ({to_int})"}

        if price_int <= 0:
            return {"ok": False, "error": f"Linje {i}: pris må være større enn 0 (fikk {price_int})"}

        if rows and frm_int != rows[-1][1]:
            return {"ok": False, "error": f"Linje {i}: gap eller overlapp i km-intervall (forrige til={rows[-1][1]}, denne fra={frm_int})"}

        rows.append((frm_int, to_int, price_int))

    if not rows:
        return {"ok": False, "error": "Filen er tom"}

    return {
        "ok": True,
        "rows": len(rows),
        "from_km": rows[0][0],
        "to_km": rows[-1][1],
        "table": rows
    }

@app.post("/ui/validate-ten")
def validate_ten(request: Request, tenFile: UploadFile = File(...)):
    require_login(request)
    global TEN_TABLE

    content = tenFile.file.read().decode("utf-8-sig")
    result = validate_ten_csv(content)

    if not result["ok"]:
        TEN_TABLE = None
        return {"ok": False, "error": result["error"]}

    TEN_TABLE = result["table"]
    return {"ok": True}


# ---------------------------------------------------------------------
# JSON byte-level utilities (avoids creating full Python object tree)
# ---------------------------------------------------------------------

def _find_json_value_end(content: bytes, start: int) -> int:
    """Return end index of the JSON value whose first byte is at start."""
    c = content[start]
    if c in (ord('{'), ord('[')):
        depth = 0
        i = start
        in_string = False
        while i < len(content):
            b = content[i]
            if in_string:
                if b == ord('\\'):
                    i += 2
                    continue
                if b == ord('"'):
                    in_string = False
            else:
                if b == ord('"'):
                    in_string = True
                elif b in (ord('{'), ord('[')):
                    depth += 1
                elif b in (ord('}'), ord(']')):
                    depth -= 1
                    if depth == 0:
                        return i + 1
            i += 1
    elif c == ord('"'):
        i = start + 1
        while i < len(content):
            b = content[i]
            if b == ord('\\'):
                i += 2
                continue
            if b == ord('"'):
                return i + 1
            i += 1
    else:
        i = start
        while i < len(content) and content[i:i+1] not in (b',', b'}', b']', b' ', b'\n', b'\r', b'\t'):
            i += 1
        return i
    return len(content)


def _find_section_span(content: bytes, key: str) -> tuple[int, int]:
    """
    Find (value_start, value_end) byte offsets for the first "key":VALUE in content.
    Raises KeyError if not found.
    """
    search = f'"{key}":'.encode()
    idx = content.find(search)
    if idx == -1:
        raise KeyError(key)
    val_start = idx + len(search)
    while val_start < len(content) and content[val_start:val_start+1] in (b' ', b'\t', b'\n', b'\r'):
        val_start += 1
    return val_start, _find_json_value_end(content, val_start)


def _apply_byte_replacements(content: bytes, replacements: list) -> bytes:
    """Apply multiple (start, end, new_bytes) replacements at once (must be non-overlapping)."""
    replacements = sorted(replacements)
    parts: list[bytes] = []
    pos = 0
    for start, end, new_bytes in replacements:
        parts.append(content[pos:start])
        parts.append(new_bytes)
        pos = end
    parts.append(content[pos:])
    return b"".join(parts)


def _append_to_json_array(content: bytes, key: str, new_items: list) -> bytes:
    """Append new_items to a JSON array identified by key in content bytes."""
    if not new_items:
        return content
    val_start, val_end = _find_section_span(content, key)
    new_bytes = (b", " + b", ".join(json.dumps(item, ensure_ascii=False).encode() for item in new_items))
    return content[:val_end - 1] + new_bytes + content[val_end - 1:]


# ---------------------------------------------------------------------
# OSDM validation
# ---------------------------------------------------------------------

def _run_osdm_validation(job_id: str, file_bytes: bytes) -> None:
    job = VALIDATION_JOBS[job_id]
    try:
        def _s():
            return io.BytesIO(file_bytes)

        # --- Check basic structure (stop at first key, no large objects) ---
        job["phase"] = "parsing"
        has_fd = False
        try:
            for prefix, event, _ in ijson.parse(_s()):
                if prefix == "fareDelivery.fareStructure" and event == "start_map":
                    has_fd = True
                    break
        except Exception:
            job.update({"status": "error", "error": "Filen er ikke gyldig JSON"})
            return

        if not has_fd:
            job.update({"status": "error", "error": "Filen mangler fareDelivery.fareStructure"})
            return

        job["phase"] = "validating"
        job["percent"] = 28

        # --- Delivery info (small object) ---
        delivery: dict = {}
        for item in ijson.items(_s(), "fareDelivery.delivery"):
            delivery = item
            break

        job["percent"] = 32

        # --- Collect constraint IDs (streaming, one pass per array) ---
        price_ids  = {item["id"] for item in ijson.items(_s(), "fareDelivery.fareStructure.prices.item") if "id" in item}
        pc_ids     = {item["id"] for item in ijson.items(_s(), "fareDelivery.fareStructure.passengerConstraints.item") if "id" in item}
        cc_ids     = {item["id"] for item in ijson.items(_s(), "fareDelivery.fareStructure.carrierConstraints.item") if "id" in item}
        bundle_ids = {item["id"] for item in ijson.items(_s(), "fareDelivery.fareStructure.fareConstraintBundles.item") if "id" in item}
        text_ids   = {item["id"] for item in ijson.items(_s(), "fareDelivery.fareStructure.texts.item") if "id" in item}
        cp_ids     = {item["id"] for item in ijson.items(_s(), "fareDelivery.fareStructure.connectionPoints.item") if "id" in item}

        rc_ids  = set()
        rc_list = []
        for item in ijson.items(_s(), "fareDelivery.fareStructure.regionalConstraints.item"):
            if "id" in item:
                rc_ids.add(item["id"])
                rc_list.append({"id": item["id"],
                                "entryConnectionPointId": item.get("entryConnectionPointId"),
                                "exitConnectionPointId":  item.get("exitConnectionPointId")})

        job["percent"] = 38

        # --- Validate fares (streaming, 1.2M+ fares processed one at a time) ---
        used_price_ids: set[str] = set()
        used_pc_ids:    set[str] = set()
        used_rc_ids:    set[str] = set()
        missing: dict[str, set[str]] = {
            "priceRef": set(), "passengerConstraintRef": set(),
            "regionalConstraintRef": set(), "carrierConstraintRef": set(),
            "bundleRef": set(), "nameRef": set(),
        }
        warnings: list[str] = []

        fare_num = 0
        for fare in ijson.items(_s(), "fareDelivery.fareStructure.fares.item"):
            fare_num += 1
            pr = fare.get("priceRef")
            if pr:
                if pr not in price_ids: missing["priceRef"].add(pr)
                else: used_price_ids.add(pr)
            pc = fare.get("passengerConstraintRef")
            if pc:
                if pc not in pc_ids: missing["passengerConstraintRef"].add(pc)
                else: used_pc_ids.add(pc)
            rc = fare.get("regionalConstraintRef")
            if rc:
                if rc not in rc_ids: missing["regionalConstraintRef"].add(rc)
                else: used_rc_ids.add(rc)
            cc = fare.get("carrierConstraintRef")
            if cc and cc not in cc_ids: missing["carrierConstraintRef"].add(cc)
            br = fare.get("bundleRef")
            if br and br not in bundle_ids: missing["bundleRef"].add(br)
            nr = fare.get("nameRef")
            if nr and nr not in text_ids: missing["nameRef"].add(nr)
            if fare_num % 50000 == 0:
                # Asymptotic progress: 38 → 72 over ~1M fares
                job["percent"] = 38 + int(34 * (1 - math.exp(-fare_num / 500000)))

        job["percent"] = 72

        # --- RC connection point check ---
        rc_bad_cps: list[str] = []
        for rc in rc_list:
            for field in ("entryConnectionPointId", "exitConnectionPointId"):
                cp_ref = rc.get(field)
                if cp_ref and cp_ref not in cp_ids:
                    rc_bad_cps.append(f"{rc['id']}: {field}={cp_ref}")

        job["percent"] = 82
        label_map = {
            "priceRef": "prices", "passengerConstraintRef": "passengerConstraints",
            "regionalConstraintRef": "regionalConstraints", "carrierConstraintRef": "carrierConstraints",
            "bundleRef": "fareConstraintBundles", "nameRef": "texts",
        }
        for field, ids in missing.items():
            if ids:
                examples = ", ".join(sorted(ids)[:5])
                more = f" (+ {len(ids) - 5} til)" if len(ids) > 5 else ""
                warnings.append(f"{len(ids)} fare(r) peker på ukjent {field} i {label_map[field]}: {examples}{more}")
        if rc_bad_cps:
            examples = "; ".join(rc_bad_cps[:3])
            more = f" (+ {len(rc_bad_cps) - 3} til)" if len(rc_bad_cps) > 3 else ""
            warnings.append(f"RC-er med ugyldig CP-referanse: {examples}{more}")

        unused_prices = price_ids - used_price_ids
        unused_pcs    = pc_ids - used_pc_ids
        unused_rcs    = rc_ids - used_rc_ids
        if unused_prices:
            warnings.append(f"{len(unused_prices)} pris(er) er definert men ikke referert av noen fare")
        if unused_pcs:
            warnings.append(f"{len(unused_pcs)} passengerConstraint(s) er definert men ikke referert av noen fare")
        if unused_rcs:
            warnings.append(f"{len(unused_rcs)} regionalConstraint(s) er definert men ikke referert av noen fare")

        job["percent"] = 92

        # --- Station count (small array, load each CP as dict) ---
        station_count = 0
        for cp in ijson.items(_s(), "fareDelivery.fareStructure.connectionPoints.item"):
            for ss in cp.get("stationSets", []):
                for s in ss if isinstance(ss, list) else [ss]:
                    if isinstance(s, dict) and s.get("codeList") == "UIC":
                        station_count += 1

        job.update({
            "status": "done", "percent": 100,
            "result": {
                "ok": True,
                "warnings": warnings,
                "deliveryId": delivery.get("deliveryId", ""),
                "fareProvider": delivery.get("fareProvider", ""),
                "fareCount": len(rc_ids),
                "priceCount": len(price_ids),
                "stationCount": station_count,
            }
        })

    except Exception as e:
        job.update({"status": "error", "error": str(e)})


@app.post("/ui/validate-osdm")
async def validate_osdm(request: Request, osdmFile: UploadFile = File(...)):
    require_login(request)
    file_bytes = await osdmFile.read()
    job_id = str(uuid.uuid4())
    VALIDATION_JOBS[job_id] = {
        "status": "running", "percent": 0, "phase": "parsing",
        "start_time": time.time(), "file_size": len(file_bytes),
        "result": None, "error": None,
        "owner": request.session.get("user_email"), "created_at": time.time(),
    }
    threading.Thread(target=_run_osdm_validation, args=(job_id, file_bytes), daemon=True).start()
    return {"jobId": job_id}


@app.get("/ui/validate-osdm/progress/{job_id}")
def validate_osdm_progress(job_id: str, request: Request):
    require_login(request)
    job = VALIDATION_JOBS.get(job_id)
    if not job:
        return {"status": "error", "error": "Jobb ikke funnet"}
    if job.get("owner") != request.session.get("user_email"):
        raise HTTPException(status_code=403, detail="Ikke tilgang")

    if job.get("phase") == "parsing":
        elapsed = time.time() - job.get("start_time", time.time())
        file_mb = job.get("file_size", 0) / (1024 * 1024)
        estimated_s = max(2.0, file_mb / 20)  # assume ~20 MB/s for json.loads
        pct = min(26, int(elapsed / estimated_s * 26))
        return {"status": "running", "percent": pct, "stage": "reading"}

    if job["status"] in ("done", "error"):
        result = {
            "status": job["status"],
            "percent": job.get("percent", 100),
            "result": job.get("result"),
            "error": job.get("error"),
        }
        VALIDATION_JOBS.pop(job_id, None)
        return result

    return {"status": "running", "percent": job.get("percent", 0), "stage": "validating"}


# ---------------------------------------------------------------------
# Distance CSV validation
# ---------------------------------------------------------------------

@app.post("/ui/validate-distances")
async def validate_distances(
    request: Request,
    distanceFile: UploadFile = File(...),
    osdmFile: UploadFile = File(...),
):
    require_login(request)

    # Parse TEN-format avstandsfil (fra_km;til_km;pris)
    try:
        csv_text = (await distanceFile.read()).decode("utf-8-sig")
    except Exception:
        return {"ok": False, "error": "Kunne ikke lese avstandsfilen"}

    ten_result = validate_ten_csv(csv_text)
    if not ten_result["ok"]:
        return {"ok": False, "error": ten_result["error"]}
    ten_table: list[tuple[int, int, int]] = ten_result["table"]  # (fra_km, til_km, pris)

    # Parse OSDM
    try:
        data = json.loads(await osdmFile.read())
    except Exception:
        return {"ok": False, "error": "OSDM-filen er ikke gyldig JSON"}

    fs = data.get("fareDelivery", {}).get("fareStructure", {})
    if not fs:
        return {"ok": False, "error": "OSDM-filen mangler fareDelivery.fareStructure"}

    # Stasjonsnavn fra stationNames
    station_names: dict[str, str] = {
        sn["code"]: sn.get("nameUtf8") or sn.get("name") or sn["code"]
        for sn in fs.get("stationNames", [])
    }

    # CP-id → beste UIC-kode og stasjonsnavn
    cp_to_uic: dict[str, str] = {}
    cp_to_name: dict[str, str] = {}
    for cp in fs.get("connectionPoints", []):
        uic_codes = [
            s["code"]
            for ss in cp.get("stationSets", [])
            for s in ss
            if s.get("codeList") == "UIC"
        ]
        best = next((c for c in uic_codes if str(c).startswith("76")), uic_codes[0] if uic_codes else None)
        if best:
            cp_to_uic[cp["id"]] = str(best)
            cp_to_name[cp["id"]] = station_names.get(str(best), str(best))

    # Samle unike avstandsverdier fra OSDM (én per RC, ignorer duplikater)
    osdm_distances: dict[int, list[str]] = {}  # km → liste med "Fra → Til"-beskrivelser
    for rc in fs.get("regionalConstraints", []):
        km = rc.get("distance")
        if km is None:
            continue
        entry = cp_to_name.get(rc.get("entryConnectionPointId", ""), rc.get("entryConnectionPointId", ""))
        exit_ = cp_to_name.get(rc.get("exitConnectionPointId", ""), rc.get("exitConnectionPointId", ""))
        osdm_distances.setdefault(km, []).append(f"{entry} → {exit_}")

    def ten_covers(km: int) -> bool:
        return any(frm <= km <= til for frm, til, _ in ten_table)

    warnings: list[str] = []
    uncovered: list[str] = []

    for km, routes in sorted(osdm_distances.items()):
        if not ten_covers(km):
            example = routes[0]
            more = f" (+ {len(routes) - 1} RC til med samme avstand)" if len(routes) > 1 else ""
            uncovered.append(f"{km} km ({example}{more})")

    if uncovered:
        cap = 15
        examples = "; ".join(uncovered[:cap])
        more = f" (+ {len(uncovered) - cap} til)" if len(uncovered) > cap else ""
        warnings.append(
            f"{len(uncovered)} avstandsverdi(er) fra OSDM dekkes ikke av TEN-tabellen: {examples}{more}"
        )

    ten_max = ten_table[-1][1]
    osdm_max = max(osdm_distances.keys()) if osdm_distances else 0
    if osdm_max > ten_max:
        warnings.append(
            f"OSDM har ruter opp til {osdm_max} km, men TEN-tabellen slutter på {ten_max} km"
        )

    return {
        "ok": True,
        "warnings": warnings,
        "tenRows": len(ten_table),
        "tenRangeKm": f"{ten_table[0][0]}–{ten_table[-1][1]}",
        "osdmDistinctDistances": len(osdm_distances),
        "osdmRcCount": sum(len(v) for v in osdm_distances.values()),
        "uncoveredCount": len(uncovered),
    }


# ---------------------------------------------------------------------
# Exchange rate
# ---------------------------------------------------------------------

@app.get("/ui/exchange-rate")
def exchange_rate(request: Request, from_: str = "EUR", to: str = "NOK"):
    require_login(request)
    try:
        resp = requests.get(
            f"https://api.frankfurter.app/latest",
            params={"from": from_.upper(), "to": to.upper()},
            timeout=5,
        )
        resp.raise_for_status()
        payload = resp.json()
        rate = payload["rates"][to.upper()]
        return {"ok": True, "from": from_.upper(), "to": to.upper(), "rate": rate, "date": payload.get("date")}
    except Exception as exc:
        logger.warning("Kunne ikke hente valutakurs %s→%s: %s", from_, to, exc)
        return {"ok": False, "error": "Kunne ikke hente valutakurs fra frankfurter.app"}


# ---------------------------------------------------------------------
# Price helpers
# ---------------------------------------------------------------------

def nok_price_from_distance(km: int):
    for frm, to, price in TEN_TABLE:
        if frm <= km <= to:
            return price
    raise ValueError(f"Ingen TEN-pris for {km} km")

def eur_amount(nok: int, rate: float):
    eur = nok * rate
    eur_rounded = math.ceil(eur / 0.20) * 0.20
    return int(eur_rounded * 100)

# ---------------------------------------------------------------------
# Generate OSDM
# ---------------------------------------------------------------------

@app.post("/ui/generate-osdm")
def generate_osdm(
    request: Request,
    exchangeRate: float = Form(...),
    validFrom: str = Form(...),
    validTo: str = Form(...),
    datasetId: str = Form(...),
    environment: str = Form(...),
    optionalDelivery: str = Form("false"),
    previousDeliveryId: str = Form(""),
):
    require_login(request)

    if TEN_TABLE is None:
        raise HTTPException(status_code=400, detail="TEN-CSV er ikke validert")

    GENERATION_PROGRESS["status"] = "running"
    GENERATION_PROGRESS["percent"] = 0

    data = json.loads(OSDM_IN.read_text(encoding="utf-8"))

    # Erstatt gammel delivery-id med ny datasetId overalt i strukturen
    old_delivery_id = data["fareDelivery"]["delivery"]["deliveryId"]
    if old_delivery_id and old_delivery_id != datasetId:
        raw = json.dumps(data)
        raw = raw.replace(f"1076_{old_delivery_id}_", f"1076_{datasetId}_")
        data = json.loads(raw)

    fs = data["fareDelivery"]["fareStructure"]

    data["fareDelivery"]["delivery"]["deliveryId"] = datasetId
    if previousDeliveryId.strip():
        data["fareDelivery"]["delivery"]["previousDeliveryId"] = previousDeliveryId.strip()
    else:
        data["fareDelivery"]["delivery"].pop("previousDeliveryId", None)
    data["fareDelivery"]["delivery"]["optionalDelivery"] = (optionalDelivery.lower() == "true")
    data["fareDelivery"]["delivery"]["usage"] = (
        "TEST_ONLY" if environment == "test" else "PRODUCTION"
    )

    # Sett datoer og utcOffset basert på Oslo-tidssone
    oslo_tz = ZoneInfo("Europe/Oslo")
    from_dt = datetime.fromisoformat(validFrom).replace(tzinfo=oslo_tz)
    until_dt = datetime.fromisoformat(validTo).replace(hour=23, minute=59, second=59, tzinfo=oslo_tz)
    utc_offset_from = int(from_dt.utcoffset().total_seconds() / 60)

    from_date = f"{validFrom}T00:00:00+0000"
    until_date = f"{validTo}T23:59:59+0000"

    for cal in fs.get("calendars", []):
        cal["fromDate"] = from_date
        cal["untilDate"] = until_date
        cal["utcOffset"] = utc_offset_from

    # Bygg eksempel-oppslag: UIC-kode -> connectionPointId
    cp_for_uic = {}
    for cp in fs["connectionPoints"]:
        for ss in cp.get("stationSets", []):
            for s in ss:
                if s.get("codeList") == "UIC":
                    cp_for_uic[s["code"]] = cp["id"]

    example_routes = [
        ("Oslo S",  "Bergen stasjon",    "7600100", "7602351"),
        ("Oslo S",  "Trondheim S",       "7600100", "7601126"),
        ("Oslo S",  "Stavanger stasjon", "7600100", "7602234"),
        ("Oslo S",  "Halden stasjon",    "7600100", "7600546"),
        ("Oslo S",  "Kornsjø grense",    "7600100", "7600551"),
    ]

    # Kategoriratio mot voksen
    # (nameRef-suffix, passengerConstraintRef-suffix, ratio)
    CATEGORY_RATIOS = [
        ("P__7",  "G__1", 1.00),   # Voksen
        ("P__34", "G__2", 0.90),   # Voksen gruppe
        ("P__11", "G__8", 0.50),   # Senior
        ("P__8",  "G__3", 0.25),   # Barn 6-17 år
        ("P__35", "G__4", 0.25),   # Barn 6-17 år gruppe
        ("P__9",  "G__6", 0.00),   # Barn 0-5 år
        ("P__36", "G__7", 0.00),   # Barn 0-5 år gruppe
        ("P__5",  "G__1", 0.50),   # FIP leisure reduction voksen
        ("P__5",  "G__3", 0.25),   # FIP leisure reduction barn
        ("P__10", "G__5", 0.50),   # Hund
    ]

    # Finn id-prefix fra eksisterende fare-id-er (f.eks. "1076_8.2_")
    sample_nr = fs["fares"][0].get("nameRef", "") if fs.get("fares") else ""
    id_prefix = "_".join(sample_nr.split("_")[:2]) + "_" if sample_nr else f"1076_{datasetId}_"

    new_prices = []
    price_index = 1
    total = len(fs["regionalConstraints"])
    examples = {}
    example_idx = 1

    # rc_id -> { (nameRef, passengerConstraintRef) -> ny price_id }
    rc_fare_price_map: dict = {}

    for idx, rc in enumerate(fs["regionalConstraints"], start=1):
        km = rc.get("distance")
        if km is None:
            continue

        nok = nok_price_from_distance(km)
        rc_fare_price_map[rc["id"]] = {}
        voksen_amount = None

        for nr_sfx, pc_sfx, ratio in CATEGORY_RATIOS:
            nr_key = id_prefix + nr_sfx
            pc_key = id_prefix + pc_sfx

            if ratio > 0:
                raw_eur = nok * exchangeRate * ratio
                cat_amount = int(math.ceil(raw_eur / 0.20) * 0.20 * 100)
            else:
                cat_amount = 0

            # Lagre voksenbeløpet for eksempelpriser
            if nr_sfx == "P__7" and pc_sfx == "G__1":
                voksen_amount = cat_amount

            price_id = f"1076_{datasetId}_I__{price_index}"
            new_prices.append({
                "id": price_id,
                "price": [{
                    "amount": cat_amount,
                    "currency": "EUR",
                    "scale": 2,
                    "vatDetails": []
                }]
            })
            rc_fare_price_map[rc["id"]][(nr_key, pc_key)] = price_id
            price_index += 1

        # Eksempelpriser bruker voksenprisen
        if voksen_amount is not None:
            for from_name, to_name, from_uic, to_uic in example_routes:
                if (
                    rc["entryConnectionPointId"] == cp_for_uic.get(from_uic)
                    and rc["exitConnectionPointId"] == cp_for_uic.get(to_uic)
                ):
                    examples[f"example_{example_idx}"] = (
                        f"{from_name} -> {to_name}: {voksen_amount / 100:.2f} EUR ({km} km)"
                    )
                    example_idx += 1

        GENERATION_PROGRESS["percent"] = int(idx / total * 100)

    fs["prices"] = new_prices

    # Oppdater priceRef i alle fares til å peke på nye price-id-er
    for fare in fs["fares"]:
        rc_ref = fare.get("regionalConstraintRef")
        nr = fare.get("nameRef")
        pc = fare.get("passengerConstraintRef")
        new_price_id = rc_fare_price_map.get(rc_ref, {}).get((nr, pc))
        if new_price_id:
            fare["priceRef"] = new_price_id

    GENERATION_PROGRESS["status"] = "done"
    GENERATION_PROGRESS["percent"] = 100

    filename = f"1076_{datasetId}_{environment}.json"
    content = json.dumps(data, indent=2, ensure_ascii=False)

    user_email = request.session.get("user_email", "")
    OSDM_STORE[user_email] = {"filename": filename, "content": content, "created_at": time.time()}

    log_event(request.session.get("user_email"), "osdm_generated", detail={
        "deliveryId": datasetId,
        "environment": environment,
        "validFrom": validFrom,
        "validTo": validTo,
        "priceCount": len(new_prices),
    })

    return {
        "step": "OSDM generation",
        "ok": True,
        "outputFile": filename,
        "summary": {
            "pricesUpdated": len(fs["prices"]),
            "exchangeRate": exchangeRate,
            "environment": environment,
            "utcOffset": utc_offset_from,
            "exampleFares": examples,
        },
    }


# ---------------------------------------------------------------------
# Progress, download & Excel export
# ---------------------------------------------------------------------

@app.get("/ui/progress")
def get_progress():
    return GENERATION_PROGRESS

@app.post("/ui/excel-from-generated")
def excel_from_generated(request: Request):
    require_login(request)
    user_email = request.session.get("user_email", "")
    osdm_entry = OSDM_STORE.get(user_email)
    if not osdm_entry:
        raise HTTPException(status_code=400, detail="Ingen generert OSDM-fil tilgjengelig")

    osdm_bytes = osdm_entry["content"].encode("utf-8") if isinstance(osdm_entry["content"], str) else osdm_entry["content"]

    # Extract delivery info for filename (quick ijson scan, no full parse)
    delivery_info: dict = {}
    try:
        for item in ijson.items(io.BytesIO(osdm_bytes), "fareDelivery.delivery"):
            delivery_info = item
            break
    except Exception:
        raise HTTPException(status_code=500, detail="Kunne ikke lese generert OSDM-fil")

    env_suffix = "test" if delivery_info.get("usage") == "TEST_ONLY" else "prod"
    gen_filename = f"{delivery_info.get('fareProvider', '')}_{delivery_info.get('deliveryId', 'osdm')}_{env_suffix}.xlsx"

    job_id = str(uuid.uuid4())
    XLSX_JOBS[job_id] = {"status": "running", "result": None, "error": None, "filename": None, "percent": 0, "rows": 0,
                         "owner": user_email, "created_at": time.time()}

    caller_email = user_email

    def run():
        try:
            xlsx_bytes, row_count = osdm_to_xlsx_bytes(osdm_bytes, job_id, XLSX_JOBS)
            XLSX_JOBS[job_id].update({"status": "done", "result": xlsx_bytes, "filename": gen_filename, "rows": row_count, "percent": 100})
            log_event(caller_email, "excel_exported", detail={"filename": gen_filename, "rows": row_count})
        except Exception as e:
            import traceback; traceback.print_exc()
            XLSX_JOBS[job_id].update({"status": "error", "error": str(e)})
            log_event(caller_email, "excel_exported", "error", {"error": str(e)})

    threading.Thread(target=run, daemon=True).start()
    return {"jobId": job_id}


@app.get("/ui/download-osdm/{filename}")
def download_osdm(filename: str, request: Request):
    require_login(request)
    user_email = request.session.get("user_email", "")
    osdm_entry = OSDM_STORE.get(user_email)
    if not osdm_entry or osdm_entry["filename"] != filename:
        raise HTTPException(status_code=404, detail="OSDM-fil finnes ikke")
    safe_name = _safe_filename(filename)
    return Response(
        content=osdm_entry["content"],
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )

# ---------------------------------------------------------------------
# Admin
# ---------------------------------------------------------------------

@app.get("/admin/list-users")
def list_users(request: Request):
    require_admin(request)
    db = SessionLocal()
    try:
        users = db.query(User).all()
        return [
            {
                "email": u.email,
                "is_admin": u.is_admin,
                "is_active": u.is_active,
                "has_logged_in": u.first_login_at is not None,
                "first_login_at": u.first_login_at.isoformat() if u.first_login_at else None,
                "must_change_password": bool(u.must_change_password),
            }
            for u in users
        ]
    finally:
        db.close()

@app.post("/admin/add-user")
def admin_add_user(request: Request, email: str = Form(...), is_admin: str = Form("false")):
    require_admin(request)
    email = email.lower()
    db = SessionLocal()
    try:
        if db.query(User).filter(User.email == email).first():
            raise HTTPException(status_code=400, detail="Bruker finnes allerede")
        password = generate_password()
        db.add(User(
            email=email,
            password_hash=hash_password(password),
            is_admin=(is_admin.lower() == "true"),
            is_active=True,
            must_change_password=True,
        ))
        db.commit()
    finally:
        db.close()
    log_event(request.session.get("user_email"), "admin_user_created", detail={
        "email": email, "is_admin": is_admin.lower() == "true",
    })
    try:
        send_welcome_email(email, password)
        return {"ok": True, "email": email, "email_sent": True}
    except Exception as exc:
        logger.error("Kunne ikke sende velkomst-e-post til %s: %s", email, exc)
        return {"ok": True, "email": email, "email_sent": False}

@app.post("/admin/reset-password")
def admin_reset_password(request: Request, email: str = Form(...)):
    require_admin(request)
    email = email.lower()
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == email).first()
        if not user or not user.is_active:
            raise HTTPException(status_code=404, detail="Bruker ikke funnet")
        new_password = generate_password()
        user.password_hash = hash_password(new_password)
        user.must_change_password = True
        db.commit()
    finally:
        db.close()
    log_event(request.session.get("user_email"), "admin_password_reset", detail={"email": email})
    try:
        send_reset_email(email, new_password)
        return {"ok": True, "email": email, "email_sent": True}
    except Exception as exc:
        logger.error("Kunne ikke sende reset-e-post til %s: %s", email, exc)
        return {"ok": True, "email": email, "email_sent": False}

@app.post("/admin/set-admin")
def admin_set_admin(request: Request, email: str = Form(...), is_admin: str = Form(...)):
    require_admin(request)
    email = email.lower()
    if email == request.session.get("user_email"):
        raise HTTPException(status_code=400, detail="Kan ikke endre din egen admin-status")
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == email).first()
        if not user:
            raise HTTPException(status_code=404, detail="Bruker ikke funnet")
        user.is_admin = (is_admin.lower() == "true")
        db.commit()
        log_event(request.session.get("user_email"), "admin_set_admin", detail={
            "email": email, "is_admin": user.is_admin,
        })
        return {"ok": True, "email": email, "is_admin": user.is_admin}
    finally:
        db.close()

@app.post("/admin/delete-user")
def delete_user(request: Request, email: str = Form(...)):
    require_admin(request)
    email = email.lower()
    if email == request.session.get("user_email"):
        raise HTTPException(status_code=400, detail="Kan ikke slette deg selv")
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == email).first()
        if not user:
            raise HTTPException(status_code=404, detail="Bruker ikke funnet")
        db.delete(user)
        db.commit()
        log_event(request.session.get("user_email"), "admin_user_deleted", detail={"email": email})
        return {"ok": True, "deleted": email}
    finally:
        db.close()

@app.get("/admin/pending-requests")
def admin_pending_requests(request: Request):
    require_admin(request)
    with SessionLocal() as db:
        rows = (
            db.query(AccessRequest)
            .filter(AccessRequest.status == "pending")
            .order_by(AccessRequest.requested_at.asc())
            .all()
        )
        return [
            {
                "email":        r.email,
                "name":         r.name,
                "org":          r.org,
                "requested_at": r.requested_at.isoformat() if r.requested_at else None,
            }
            for r in rows
        ]


@app.post("/admin/approve-request")
def admin_approve_request(request: Request, email: str = Form(...)):
    require_admin(request)
    email = email.lower()
    with SessionLocal() as db:
        req = db.query(AccessRequest).filter(AccessRequest.email == email).first()
        if not req or req.status != "pending":
            raise HTTPException(status_code=404, detail="Forespørsel ikke funnet")
        existing = db.query(User).filter(User.email == email).first()
        if existing:
            req.status = "approved"
            db.commit()
            return {"ok": True, "email": email, "already_existed": True}
        password = generate_password()
        user = User(
            email=email,
            password_hash=hash_password(password),
            is_admin=False,
            is_active=True,
            must_change_password=True,
            first_name=req.first_name or "",
            last_name=req.last_name or "",
        )
        db.add(user)
        req.status = "approved"
        db.commit()
    email_sent = True
    try:
        send_welcome_email(email, password)
    except Exception as exc:
        logger.warning("Kunne ikke sende velkomstepost: %s", exc)
        email_sent = False
    log_event(request.session.get("user_email"), "admin_request_approved", detail={"email": email})
    return {"ok": True, "email": email, "email_sent": email_sent}


@app.post("/admin/reject-request")
def admin_reject_request(request: Request, email: str = Form(...)):
    require_admin(request)
    email = email.lower()
    with SessionLocal() as db:
        req = db.query(AccessRequest).filter(AccessRequest.email == email).first()
        if not req:
            raise HTTPException(status_code=404, detail="Forespørsel ikke funnet")
        req.status = "rejected"
        db.commit()
    log_event(request.session.get("user_email"), "admin_request_rejected", detail={"email": email})
    return {"ok": True, "email": email}


@app.get("/admin/login-log")
def admin_login_log(
    request: Request,
    search: str = "",
    date_from: str = "",
    date_to: str = "",
    page: int = 1,
    page_size: int = 25,
):
    require_admin(request)
    db = SessionLocal()
    try:
        q = db.query(LoginLog)
        if search:
            q = q.filter(LoginLog.email.ilike(f"%{search}%"))
        if date_from:
            q = q.filter(LoginLog.logged_at >= date_from)
        if date_to:
            dt_to = datetime.fromisoformat(date_to) + timedelta(days=1)
            q = q.filter(LoginLog.logged_at < dt_to)
        total = q.count()
        entries = (
            q.order_by(LoginLog.logged_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "entries": [
                {
                    "email": e.email,
                    "logged_at": e.logged_at.isoformat() if e.logged_at else None,
                    "ip_address": e.ip_address or "—",
                    "success": bool(e.success) if e.success is not None else True,
                }
                for e in entries
            ],
        }
    finally:
        db.close()

# ---------------------------------------------------------------------
# OSDM til CSV
# ---------------------------------------------------------------------

def suffix(id_str: str) -> str:
    """Hent suffix etter siste '__', f.eks '1076_7.0_P__7' -> 'P__7'"""
    parts = id_str.split("_")
    # finn P__, G__, osv
    for i, p in enumerate(parts):
        if p in ("P", "G", "I", "K", "E", "S", "C", "M", "Q", "T", "D"):
            return p + "__" + parts[i + 1]
    return id_str


from backend.rics_codes import RICS_CODES as RICS_CARRIER_NAMES, RICS_COUNTRIES


def osdm_to_xlsx_bytes(file_bytes: bytes, job_id: str = None, jobs: dict = None) -> tuple[bytes, int]:
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timezone, timedelta

    def set_progress(percent: int):
        if job_id and jobs and job_id in jobs:
            jobs[job_id]["percent"] = percent

    def _s():
        return _io.BytesIO(file_bytes)

    # --- Build lookup maps via ijson (no full parse) ---
    delivery: dict = {}
    for item in ijson.items(_s(), "fareDelivery.delivery"):
        delivery = item
        break
    delivery_id = delivery.get("deliveryId", "")

    text_map: dict[str, str] = {}
    for t in ijson.items(_s(), "fareDelivery.fareStructure.texts.item"):
        text_map[t["id"]] = t.get("textUtf8", t.get("text", ""))

    price_map: dict[str, int] = {}
    for p in ijson.items(_s(), "fareDelivery.fareStructure.prices.item"):
        if p.get("price"):
            price_map[p["id"]] = p["price"][0]["amount"]

    station_map: dict[str, str] = {}
    for sn in ijson.items(_s(), "fareDelivery.fareStructure.stationNames.item"):
        code = sn.get("code") or sn.get("uicCode")
        if code:
            station_map[str(code)] = sn.get("nameUtf8", sn.get("name", ""))

    service_class_map: dict[str, str] = {}
    for sc in ijson.items(_s(), "fareDelivery.fareStructure.serviceClassDefinitions.item"):
        service_class_map[sc["id"]] = text_map.get(sc.get("textRef", ""), sc["id"])

    pc_map: dict[str, dict] = {}
    for pc in ijson.items(_s(), "fareDelivery.fareStructure.passengerConstraints.item"):
        pc_map[pc["id"]] = pc

    svc_constraint_map: dict[str, str] = {}
    for sc in ijson.items(_s(), "fareDelivery.fareStructure.serviceConstraints.item"):
        svc_constraint_map[sc["id"]] = text_map.get(sc.get("textRef", ""), sc["id"])

    # Reduction-oppslag: id -> korteste representative kortnavn
    reduction_map: dict[str, str] = {}
    for rc in ijson.items(_s(), "fareDelivery.fareStructure.reductionConstraints.item"):
        cards = rc.get("requiredCards", [])
        if cards:
            name = min(cards, key=lambda c: len(c.get("cardName", "")))
            reduction_map[rc["id"]] = name.get("cardName", rc["id"])

    set_progress(10)

    PASSENGER_TYPE_LABELS = {
        "ADULT": "Adult", "CHILD": "Child", "YOUNG_CHILD": "Young child",
        "SENIOR": "Senior", "DOG": "Dog", "INFANT": "Infant",
    }

    # CP-id -> beste UIC-kode og stasjonsnavn
    cp_to_uic: dict[str, str] = {}
    cp_to_name: dict[str, str] = {}
    for cp in ijson.items(_s(), "fareDelivery.fareStructure.connectionPoints.item"):
        best_code = None
        all_codes = []
        for ss in cp.get("stationSets", []):
            for s in ss:
                if s.get("codeList") == "UIC":
                    code = s["code"]
                    all_codes.append(str(code))
                    if best_code is None or s.get("country") == "NO":
                        best_code = str(code)
        best_name = None
        for code in all_codes:
            name = station_map.get(code)
            if name:
                best_name = name
                break
        cp_to_uic[cp["id"]] = best_code or ""
        cp_to_name[cp["id"]] = best_name or best_code or ""

    rc_map: dict[str, dict] = {}
    for rc in ijson.items(_s(), "fareDelivery.fareStructure.regionalConstraints.item"):
        rc_map[rc["id"]] = rc

    cal: dict = {}
    for item in ijson.items(_s(), "fareDelivery.fareStructure.calendars.item"):
        cal = item
        break

    carriers_set: set[str] = set()
    for cc in ijson.items(_s(), "fareDelivery.fareStructure.carrierConstraints.item"):
        for c in cc.get("includedCarrier", []):
            carriers_set.add(c)

    set_progress(20)

    # Single pass through fares: build seen_categories AND rc_prices simultaneously
    seen_categories: dict = {}
    category_order: list = []
    rc_prices: dict = {}
    fare_count = 0

    for fare in ijson.items(_s(), "fareDelivery.fareStructure.fares.item"):
        fare_count += 1
        nr = fare.get("nameRef", "")
        pc_ref = fare.get("passengerConstraintRef", "")
        sc = fare.get("serviceClassRef", "")
        red = fare.get("reductionConstraintRef", "")
        svc = fare.get("serviceConstraintRef", "")
        key = (nr, pc_ref, sc, red, svc)

        # Category discovery (pass 1 logic)
        if key not in seen_categories:
            parts = [text_map.get(nr, nr)]
            sc_name = service_class_map.get(sc, "")
            if sc_name:
                parts.append(sc_name)
            red_name = reduction_map.get(red, "")
            if red_name:
                parts.append(red_name)
            svc_name = svc_constraint_map.get(svc, "")
            if svc_name:
                parts.append(svc_name)
            seen_categories[key] = " ".join(parts)
            category_order.append(key)

        # Price collection (pass 2 logic)
        rc_ref = fare.get("regionalConstraintRef")
        price_ref = fare.get("priceRef")
        amount = price_map.get(price_ref)
        if amount is not None and rc_ref:
            rc_prices.setdefault(rc_ref, {})[key] = amount

        if fare_count % 100000 == 0:
            set_progress(20 + min(40, int(fare_count / 10000)))

    set_progress(60)

    # Pass 2 (in-memory): legg til passasjertype kun for duplikater
    name_counts: dict = {}
    for name in seen_categories.values():
        name_counts[name] = name_counts.get(name, 0) + 1
    duplicates = {name for name, count in name_counts.items() if count > 1}

    for key in category_order:
        if seen_categories[key] in duplicates:
            nr, pc_ref, sc, red, svc = key
            pc = pc_map.get(pc_ref, {})
            ptype = PASSENGER_TYPE_LABELS.get(pc.get("passengerType", ""), "")
            fare_name = text_map.get(nr, nr)
            sc_name = service_class_map.get(sc, "")
            red_name = reduction_map.get(red, "")
            svc_name = svc_constraint_map.get(svc, "")
            parts = [fare_name]
            if ptype:
                parts.append(ptype)
            if sc_name:
                parts.append(sc_name)
            if red_name:
                parts.append(red_name)
            if svc_name:
                parts.append(svc_name)
            seen_categories[key] = " ".join(parts)

    set_progress(70)

    # Bygg rader: slå sammen priser fra alle RC-er med samme stasjonsparparet
    pair_data: dict = {}  # sortert UIC-par -> {"entry_cp", "exit_cp", "rc", "prices"}

    for rc_id, prices in rc_prices.items():
        rc = rc_map.get(rc_id)
        if not rc:
            continue
        entry_cp = rc["entryConnectionPointId"]
        exit_cp = rc["exitConnectionPointId"]
        entry_uic = cp_to_uic.get(entry_cp, "")
        exit_uic = cp_to_uic.get(exit_cp, "")

        pair = tuple(sorted([entry_uic, exit_uic]))
        if pair not in pair_data:
            pair_data[pair] = {
                "entry_cp": entry_cp,
                "exit_cp": exit_cp,
                "rc": rc,
                "prices": {},
            }
        pair_data[pair]["prices"].update(prices)

    rows = []
    for pdata in pair_data.values():
        entry_cp = pdata["entry_cp"]
        exit_cp = pdata["exit_cp"]
        rc = pdata["rc"]
        prices = pdata["prices"]
        row = {
            "From UIC": cp_to_uic.get(entry_cp, ""),
            "From station": cp_to_name.get(entry_cp, ""),
            "To UIC": cp_to_uic.get(exit_cp, ""),
            "To station": cp_to_name.get(exit_cp, ""),
            "Km": rc.get("distance", ""),
        }
        for key in category_order:
            amount = prices.get(key)
            row[seen_categories[key]] = round(amount / 100, 2) if amount is not None else None
        rows.append(row)

    rows.sort(key=lambda r: r["From station"])

    set_progress(80)

    # Bygg XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = f"Priser {delivery_id}"

    fieldnames = (
        ["From UIC", "From station", "To UIC", "To station", "Km"]
        + [seen_categories[key] for key in category_order]
    )
    num_cols = len(fieldnames)
    last_col = get_column_letter(num_cols)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="0066CC")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="E8F1FB")
    price_col_start = 6

    # --- Metadata-seksjon øverst ---
    utc_offset = cal.get("utcOffset", 0)
    tz = timezone(timedelta(minutes=utc_offset))

    def parse_osdm_date(dt_str):
        if not dt_str:
            return ""
        try:
            dt = datetime.fromisoformat(dt_str.replace("+0000", "+00:00"))
            return dt.astimezone(tz).strftime("%Y-%m-%d")
        except Exception:
            return dt_str[:10]

    valid_from  = parse_osdm_date(cal.get("fromDate", ""))
    valid_until = parse_osdm_date(cal.get("untilDate", ""))
    carriers = sorted(carriers_set)
    carriers_display = ", ".join(
        f"{RICS_CARRIER_NAMES[c]} ({c})" if c in RICS_CARRIER_NAMES else c
        for c in carriers
    )
    usage = delivery.get("usage", "")

    thin = Side(style="thin", color="B0C4DE")
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    label_fill   = PatternFill("solid", start_color="D0E4F7")
    value_fill   = PatternFill("solid", start_color="F5F9FF")
    label_font   = Font(name="Arial", size=10, bold=True, color="1A3A5C")
    value_font   = Font(name="Arial", size=10, color="1A1A1A")
    label_align  = Alignment(horizontal="left", vertical="center", indent=1)

    # Rad 1: tittelbar
    ws.merge_cells(f"A1:{last_col}1")
    tc = ws["A1"]
    tc.value = (
        f"OSDM Fare Delivery  —  "
        f"Provider {delivery.get('fareProvider', '')}  /  "
        f"Delivery {delivery.get('deliveryId', '')}"
    )
    tc.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    tc.fill      = PatternFill("solid", start_color="003A7A")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Rader 2-6: to label+verdi-par per rad
    info_rows = [
        ("Fare Provider", delivery.get("fareProvider", ""),
         "Valid from",    valid_from),
        ("Delivery ID",   delivery.get("deliveryId", ""),
         "Valid until",   valid_until),
        ("OSDM version",  delivery.get("version", ""),
         "Carriers",      carriers_display),
        ("Usage",         usage,
         "Route pairs",   str(len(pair_data))),
        ("Optional",      "Yes" if delivery.get("optionalDelivery") else "No",
         "Fares",         f"{fare_count:,}"),
    ]

    for r_off, (lbl1, val1, lbl2, val2) in enumerate(info_rows, start=2):
        for col, text, is_label in (
            (1, lbl1, True), (2, val1, False),
            (3, lbl2, True), (4, val2, False),
        ):
            cell = ws.cell(row=r_off, column=col, value=text)
            cell.font      = label_font if is_label else value_font
            cell.fill      = label_fill if is_label else value_fill
            cell.border    = cell_border
            cell.alignment = label_align
        # Fargelegg usage-verdien
        if lbl1 == "Usage":
            uc = ws.cell(row=r_off, column=2)
            if val1 == "PRODUCTION":
                uc.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                uc.fill = PatternFill("solid", start_color="2E7D32")
            elif "TEST" in val1:
                uc.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                uc.fill = PatternFill("solid", start_color="C65000")
        ws.row_dimensions[r_off].height = 18

    # Rad 7: tom separator — pristabell starter på rad 8
    TABLE_HEADER_ROW = 8
    TABLE_DATA_START = 9

    # --- Pristabell-header (rad 8) ---
    for col_idx, col_name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=TABLE_HEADER_ROW, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # --- Pristabell-data (rad 9+) ---
    for row_idx, row in enumerate(rows, start=TABLE_DATA_START):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(fieldnames, start=1):
            val = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            if fill:
                cell.fill = fill
            if col_idx == 5:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx >= price_col_start:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 8
    for col_idx in range(price_col_start, num_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    ws.freeze_panes = f"A{TABLE_DATA_START}"
    ws.auto_filter.ref = f"A{TABLE_HEADER_ROW}:{last_col}{TABLE_HEADER_ROW + len(rows)}"
    ws.row_dimensions[TABLE_HEADER_ROW].height = 30

    set_progress(95)
    row_count = len(rows)
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), row_count


@app.post("/frontend/osdm-to-csv")
async def osdm_to_csv(
    request: Request,
    osdmFile: UploadFile = File(...),
):
    require_login(request)

    content = await osdmFile.read()
    print(f"Mottok fil: {osdmFile.filename}, størrelse: {len(content)} bytes")

    # Quick structure check with ijson (no full parse)
    has_fd = False
    try:
        for prefix, event, _ in ijson.parse(io.BytesIO(content)):
            if prefix == "fareDelivery.fareStructure" and event == "start_map":
                has_fd = True
                break
    except Exception as e:
        print(f"JSON-feil: {e}")
        raise HTTPException(status_code=400, detail="Ugyldig JSON-fil")
    if not has_fd:
        print("Struktursjekk feilet")
        raise HTTPException(status_code=400, detail="Filen ser ikke ut som en gyldig OSDM fareDelivery")

    # Extract delivery info for filename (before starting thread)
    delivery_info: dict = {}
    for item in ijson.items(io.BytesIO(content), "fareDelivery.delivery"):
        delivery_info = item
        break
    fare_provider = delivery_info.get("fareProvider", "")
    delivery_id_str = delivery_info.get("deliveryId", "osdm")
    usage_str = delivery_info.get("usage", "")
    env_suffix = "test" if usage_str == "TEST_ONLY" else "prod"
    filename = f"{fare_provider}_{delivery_id_str}_{env_suffix}.xlsx"

    job_id = str(uuid.uuid4())
    print(f"Starter jobb: {job_id}")
    caller_email = request.session.get("user_email")
    XLSX_JOBS[job_id] = {"status": "running", "result": None, "error": None, "filename": None, "percent": 0, "rows": 0,
                         "owner": caller_email, "created_at": time.time()}

    def run():
        try:
            print(f"run() starter for jobb {job_id}")
            xlsx_bytes, row_count = osdm_to_xlsx_bytes(content, job_id, XLSX_JOBS)
            XLSX_JOBS[job_id]["result"] = xlsx_bytes
            XLSX_JOBS[job_id]["rows"] = row_count
            XLSX_JOBS[job_id]["filename"] = filename
            XLSX_JOBS[job_id]["status"] = "done"
            XLSX_JOBS[job_id]["percent"] = 100
            log_event(caller_email, "excel_exported", detail={"filename": filename, "rows": row_count})
        except Exception as e:
            print(f"run() feilet: {e}")
            import traceback
            traceback.print_exc()
            XLSX_JOBS[job_id]["status"] = "error"
            XLSX_JOBS[job_id]["error"] = str(e)
            log_event(caller_email, "excel_exported", "error", {"error": str(e)})

    threading.Thread(target=run, daemon=True).start()
    print(f"Returnerer jobId: {job_id}")
    return {"jobId": job_id}

@app.get("/frontend/osdm-to-csv-status/{job_id}")
def osdm_to_csv_status(job_id: str, request: Request):
    require_login(request)
    job = XLSX_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Jobb ikke funnet")
    if job.get("owner") != request.session.get("user_email"):
        raise HTTPException(status_code=403, detail="Ikke tilgang")
    return {
        "status": job["status"],
        "percent": job.get("percent", 0),
        "error": job.get("error"),
        "filename": job.get("filename"),
        "rows": job.get("rows", 0),
    }

@app.get("/frontend/osdm-to-csv-download/{job_id}")
def osdm_to_csv_download(job_id: str, request: Request):
    from fastapi.responses import StreamingResponse
    import io as _io
    require_login(request)
    job = XLSX_JOBS.get(job_id)
    if not job or job["status"] != "done":
        raise HTTPException(status_code=404, detail="Fil ikke klar")
    if job.get("owner") != request.session.get("user_email"):
        raise HTTPException(status_code=403, detail="Ikke tilgang")
    xlsx_bytes = job["result"]
    filename = _safe_filename(job["filename"] or "osdm.xlsx")
    del XLSX_JOBS[job_id]
    return StreamingResponse(
        _io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@app.get("/admin", response_class=HTMLResponse)
def admin_redirect(request: Request):
    if "user_email" not in request.session or not _check_is_admin_from_db(request):
        return RedirectResponse("/", status_code=302)
    return RedirectResponse("/admin/users", status_code=302)

@app.get("/admin/users", response_class=HTMLResponse)
def admin_users_page(request: Request):
    if "user_email" not in request.session or not _check_is_admin_from_db(request):
        return RedirectResponse("/", status_code=302)
    html = Path("frontend/admin.html").read_text(encoding="utf-8")
    html = html.replace("</head>", "<script>window.IS_ADMIN = true;</script></head>")
    return HTMLResponse(html)

@app.get("/admin/log", response_class=HTMLResponse)
def admin_log_page(request: Request):
    if "user_email" not in request.session or not _check_is_admin_from_db(request):
        return RedirectResponse("/", status_code=302)
    html = Path("frontend/admin-log.html").read_text(encoding="utf-8")
    html = html.replace("</head>", "<script>window.IS_ADMIN = true;</script></head>")
    return HTMLResponse(html)

@app.get("/admin/presentation", response_class=HTMLResponse)
def admin_presentation_page(request: Request):
    if "user_email" not in request.session or not _check_is_admin_from_db(request):
        return RedirectResponse("/", status_code=302)
    html = Path("frontend/presentation.html").read_text(encoding="utf-8")
    html = html.replace("</head>", "<script>window.IS_ADMIN = true;</script></head>")
    return HTMLResponse(html)

@app.get("/admin/event-log")
def admin_event_log(
    request: Request,
    search: str = "",
    date_from: str = "",
    date_to: str = "",
    event_type: str = "",
    status: str = "",
    page: int = 1,
    page_size: int = 50,
):
    require_admin(request)
    db = SessionLocal()
    try:
        q = db.query(EventLog)
        if search:
            q = q.filter(EventLog.user_email.ilike(f"%{search}%"))
        if date_from:
            q = q.filter(EventLog.logged_at >= date_from)
        if date_to:
            dt_to = datetime.fromisoformat(date_to) + timedelta(days=1)
            q = q.filter(EventLog.logged_at < dt_to)
        if event_type:
            q = q.filter(EventLog.event_type == event_type)
        if status:
            q = q.filter(EventLog.status == status)
        total = q.count()
        entries = (
            q.order_by(EventLog.logged_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "entries": [
                {
                    "id": e.id,
                    "logged_at": e.logged_at.isoformat() if e.logged_at else None,
                    "user_email": e.user_email or "—",
                    "event_type": e.event_type,
                    "status": e.status or "ok",
                    "detail": json.loads(e.detail) if e.detail else {},
                }
                for e in entries
            ],
        }
    finally:
        db.close()

@app.get("/kontakt", response_class=HTMLResponse)
@app.head("/kontakt")
def kontakt_page(request: Request):
    if "user_email" not in request.session:
        return HTMLResponse(Path("frontend/login.html").read_text(encoding="utf-8"))
    is_admin   = bool(request.session.get("is_admin"))
    user_email = request.session.get("user_email", "")
    with SessionLocal() as db:
        user = db.query(User).filter(User.email == user_email).first()
        first = (user.first_name or "").strip() if user else ""
        last  = (user.last_name  or "").strip() if user else ""
    user_name = f"{first} {last}".strip()
    html = Path("frontend/contact.html").read_text(encoding="utf-8")
    html = html.replace(
        "</head>",
        f"<script>window.IS_ADMIN = {str(is_admin).lower()}; window.USER_EMAIL = {json.dumps(user_email)}; window.USER_NAME = {json.dumps(user_name)};</script></head>"
    )
    return HTMLResponse(html)

@app.get("/endre-passord")
def endre_passord_redirect():
    return RedirectResponse("/min-konto", status_code=301)

@app.get("/min-konto", response_class=HTMLResponse)
@app.head("/min-konto")
def min_konto_page(request: Request):
    if "user_email" not in request.session:
        return HTMLResponse(Path("frontend/login.html").read_text(encoding="utf-8"))
    is_admin = bool(request.session.get("is_admin"))
    html = Path("frontend/min-konto.html").read_text(encoding="utf-8")
    html = html.replace(
        "</head>",
        f"<script>window.IS_ADMIN = {str(is_admin).lower()};</script></head>"
    )
    return HTMLResponse(html)

@app.get("/account")
def get_account(request: Request):
    require_login(request)
    user_email = request.session.get("user_email")
    with SessionLocal() as db:
        user = db.query(User).filter(User.email == user_email).first()
        if not user:
            raise HTTPException(status_code=404)
        return {
            "email": user.email,
            "first_name": user.first_name or "",
            "last_name": user.last_name or "",
            "is_admin": user.is_admin,
        }

@app.post("/account/name")
def update_account_name(request: Request, first_name: str = Form(...), last_name: str = Form(...)):
    require_login(request)
    user_email = request.session.get("user_email")
    first_name = first_name.strip()
    last_name = last_name.strip()
    with SessionLocal() as db:
        user = db.query(User).filter(User.email == user_email).first()
        if not user:
            raise HTTPException(status_code=404)
        user.first_name = first_name
        user.last_name = last_name
        db.commit()
    log_event(user_email, "account_name_updated")
    return {"ok": True}

@app.delete("/account")
def delete_account(request: Request):
    require_login(request)
    user_email = request.session.get("user_email")
    with SessionLocal() as db:
        user = db.query(User).filter(User.email == user_email).first()
        if not user:
            raise HTTPException(status_code=404)
        if user.is_admin:
            raise HTTPException(status_code=403, detail="Admins kan ikke slette sin egen konto")
        db.query(EventLog).filter(EventLog.user_email == user_email).delete()
        db.query(LoginLog).filter(LoginLog.email == user_email).delete()
        db.query(PasswordResetToken).filter(PasswordResetToken.email == user_email).delete()
        db.delete(user)
        db.commit()
    log_event(None, "account_deleted", detail={"email": user_email})
    request.session.clear()
    return {"ok": True}

@app.post("/contact")
def contact(
    name: str = Form(...),
    email: str = Form(...),
    message: str = Form(...),
):
    try:
        send_contact_email(name, email, message)
        log_event(email, "contact_sent", detail={"name": name})
        return {"ok": True}
    except Exception as exc:
        logger.error("Kunne ikke sende kontakt-e-post: %s", exc)
        log_event(email, "contact_sent", "error", {"name": name})
        return {"ok": False}

@app.post("/request-access")
def request_access(
    request: Request,
    first_name: str = Form(""),
    last_name: str = Form(""),
    email: str = Form(...),
    org: str = Form(...),
    website: str = Form(""),
):
    if website:
        return {"ok": True}
    ip = _get_client_ip(request)
    _check_access_rate_limit(ip)
    email_clean      = email.strip().lower()
    first_name_clean = first_name.strip()
    last_name_clean  = last_name.strip()
    name_clean       = f"{first_name_clean} {last_name_clean}".strip()
    org_clean        = org.strip()
    with SessionLocal() as db:
        existing_user = db.query(User).filter(User.email == email_clean).first()
        if existing_user:
            return {"ok": True}
        existing_req = db.query(AccessRequest).filter(AccessRequest.email == email_clean).first()
        if existing_req:
            existing_req.name       = name_clean
            existing_req.first_name = first_name_clean
            existing_req.last_name  = last_name_clean
            existing_req.org        = org_clean
            existing_req.status     = "pending"
            existing_req.requested_at = datetime.now(timezone.utc)
        else:
            db.add(AccessRequest(
                email=email_clean, name=name_clean,
                first_name=first_name_clean, last_name=last_name_clean,
                org=org_clean,
            ))
        db.commit()
    with SessionLocal() as db:
        admin_emails = [u.email for u in db.query(User).filter(User.is_admin == True, User.is_active == True).all()]
    try:
        send_access_request_email(name_clean, email_clean, org_clean, recipients=admin_emails or None)
    except Exception as exc:
        logger.warning("Kunne ikke sende tilgangsforespørsel-epost: %s", exc)
    log_event(None, "access_request", detail={"name": name_clean, "email": email_clean, "org": org_clean})
    return {"ok": True}

@app.get("/price-adjust", response_class=HTMLResponse)
@app.head("/price-adjust")
def price_adjust_page(request: Request):
    if "user_email" not in request.session:
        return HTMLResponse(Path("frontend/login.html").read_text(encoding="utf-8"))
    is_admin = bool(request.session.get("is_admin"))
    html = Path("frontend/price-adjust.html").read_text(encoding="utf-8")
    html = html.replace(
        "</head>",
        f"<script>window.IS_ADMIN = {str(is_admin).lower()};</script></head>"
    )
    return HTMLResponse(html)


@app.post("/price-adjust")
async def price_adjust(
    request: Request,
    osdm_file: UploadFile = File(...),
    pct: float = Form(...),
    delivery_id: str = Form(...),
    previous_delivery_id: str = Form(""),
    environment: str = Form("prod"),
    optional_delivery: str = Form("false"),
    valid_from: str = Form(...),
    valid_to: str = Form(...),
):
    require_login(request)
    factor = 1 + pct / 100

    try:
        content = await osdm_file.read()
    except Exception:
        raise HTTPException(status_code=400, detail="Filen kunne ikke leses")

    # Quick structure check
    has_fs = False
    try:
        for prefix, event, _ in ijson.parse(io.BytesIO(content)):
            if prefix == "fareDelivery.fareStructure" and event == "start_map":
                has_fs = True
                break
    except Exception:
        raise HTTPException(status_code=400, detail="Filen er ikke gyldig JSON")
    if not has_fs:
        raise HTTPException(status_code=400, detail="Mangler fareDelivery.fareStructure")

    # Load prices list via ijson
    prices: list = list(ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.prices.item"))
    if not prices:
        raise HTTPException(status_code=400, detail="Mangler felt i OSDM-strukturen: 'prices'")

    # Indeks: price_id → list-posisjon
    price_idx = {p["id"]: i for i, p in enumerate(prices)}

    def get_eur(price_id: str) -> float:
        p = prices[price_idx[price_id]]
        entry = p["price"][0]
        scale = entry.get("scale", 2)
        return entry["amount"] / (10 ** scale)

    # Grupper farer etter (RC, carrier, bundle) → finn voksenpris per gruppe
    groups: dict[tuple, list] = defaultdict(list)
    for fare in ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.fares.item"):
        key = (
            fare.get("regionalConstraintRef"),
            fare.get("carrierConstraintRef"),
            fare.get("fareConstraintBundleRef"),
        )
        price_ref = fare.get("priceRef")
        if key[0] and price_ref and price_ref in price_idx:
            groups[key].append((price_ref, get_eur(price_ref)))

    # Beregn nye priser i EUR, lagre som {price_id: ny_eur}
    new_eur_amounts: dict[str, float] = {}
    for group_fares in groups.values():
        if not group_fares:
            continue
        max_eur = max(eur for _, eur in group_fares)
        if max_eur <= 0:
            continue
        new_adult_eur = math.ceil(max_eur * factor / 0.20) * 0.20
        for price_ref, eur in group_fares:
            if eur <= 0:
                new_eur_amounts[price_ref] = 0.0
            elif eur == max_eur:
                new_eur_amounts[price_ref] = new_adult_eur
            else:
                ratio = eur / max_eur
                new_eur_amounts[price_ref] = math.ceil(new_adult_eur * ratio / 0.20) * 0.20

    # Oppdater prices-lista (skriver tilbake som skalert heltall)
    for price in prices:
        if price["id"] in new_eur_amounts:
            entry = price["price"][0]
            scale = entry.get("scale", 2)
            entry["amount"] = int(round(new_eur_amounts[price["id"]] * (10 ** scale)))

    # Load delivery and calendars via ijson
    delivery: dict = {}
    for item in ijson.items(io.BytesIO(content), "fareDelivery.delivery"):
        delivery = item
        break
    calendars: list = list(ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.calendars.item"))

    old_id = delivery.get("deliveryId", "")

    # Oppdater delivery-felt
    delivery["deliveryId"] = delivery_id
    if previous_delivery_id.strip():
        delivery["previousDeliveryId"] = previous_delivery_id.strip()
    else:
        delivery.pop("previousDeliveryId", None)
    delivery["optionalDelivery"] = (optional_delivery.lower() == "true")
    delivery["usage"] = "TEST_ONLY" if environment == "test" else "PRODUCTION"

    oslo_tz = ZoneInfo("Europe/Oslo")
    from_dt = datetime.fromisoformat(valid_from).replace(tzinfo=oslo_tz)
    utc_offset = int(from_dt.utcoffset().total_seconds() / 60)
    from_date = f"{valid_from}T00:00:00+0000"
    until_date = f"{valid_to}T23:59:59+0000"
    for cal in calendars:
        cal["fromDate"] = from_date
        cal["untilDate"] = until_date
        cal["utcOffset"] = utc_offset

    # If deliveryId changed, do byte-level string-replace first
    if old_id and old_id != delivery_id:
        content = content.replace(f"_{old_id}_".encode(), f"_{delivery_id}_".encode())

    # Find byte spans and replace sections
    try:
        delivery_span = _find_section_span(content, "delivery")
        prices_span   = _find_section_span(content, "prices")
        calendars_span = _find_section_span(content, "calendars")
    except KeyError as e:
        raise HTTPException(status_code=400, detail=f"Mangler felt i OSDM-strukturen: {e}")

    result_bytes = _apply_byte_replacements(content, [
        (delivery_span[0],  delivery_span[1],  json.dumps(delivery,  ensure_ascii=False).encode("utf-8")),
        (prices_span[0],    prices_span[1],    json.dumps(prices,    ensure_ascii=False).encode("utf-8")),
        (calendars_span[0], calendars_span[1], json.dumps(calendars, ensure_ascii=False).encode("utf-8")),
    ])

    fare_provider = delivery.get("fareProvider", "")
    env_suffix = "test" if environment == "test" else "prod"
    filename = _safe_filename(f"{fare_provider}_{delivery_id}_{env_suffix}.json")

    log_event(
        request.session.get("user_email"), "price_adjust",
        detail={"factor": factor, "filename": osdm_file.filename, "prices_updated": len(new_eur_amounts), "delivery_id": delivery_id},
    )

    return Response(
        content=result_bytes,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/fare-discount/rics")
def fare_discount_rics(request: Request):
    require_login(request)
    return [
        {"code": code, "name": name, "country": RICS_COUNTRIES.get(code, "")}
        for code, name in sorted(RICS_CARRIER_NAMES.items(), key=lambda x: x[1])
    ]


@app.get("/fare-discount", response_class=HTMLResponse)
@app.head("/fare-discount")
def fare_discount_page(request: Request):
    if "user_email" not in request.session:
        return HTMLResponse(Path("frontend/login.html").read_text(encoding="utf-8"))
    is_admin = bool(request.session.get("is_admin"))
    html = Path("frontend/fare-discount.html").read_text(encoding="utf-8")
    html = html.replace(
        "</head>",
        f"<script>window.IS_ADMIN = {str(is_admin).lower()};</script></head>"
    )
    return HTMLResponse(html)

def _run_fare_discount_parse(job_id: str, file_bytes: bytes) -> None:
    job = PARSE_JOBS[job_id]
    try:
        def _s():
            return io.BytesIO(file_bytes)

        job["phase"] = "parsing"
        has_fd = False
        try:
            for prefix, event, _ in ijson.parse(_s()):
                if prefix == "fareDelivery.fareStructure" and event == "start_map":
                    has_fd = True
                    break
        except Exception:
            job.update({"status": "error", "error": "Filen er ikke gyldig JSON"})
            return
        if not has_fd:
            job.update({"status": "error", "error": "Filen mangler fareDelivery.fareStructure"})
            return

        job["phase"] = "validating"
        job["percent"] = 20

        delivery: dict = {}
        for item in ijson.items(_s(), "fareDelivery.delivery"):
            delivery = item
            break

        job["percent"] = 25

        texts_map: dict = {}
        for t in ijson.items(_s(), "fareDelivery.fareStructure.texts.item"):
            texts_map[t["id"]] = t

        job["percent"] = 35

        uic_to_name: dict[str, str] = {}
        for sn in ijson.items(_s(), "fareDelivery.fareStructure.stationNames.item"):
            code = sn.get("code") or sn.get("uicCode")
            name = sn.get("nameUtf8") or sn.get("name") or code
            if code:
                uic_to_name[str(code)] = name

        job["percent"] = 50

        stations: list[dict] = []
        seen: set[str] = set()
        for cp in ijson.items(_s(), "fareDelivery.fareStructure.connectionPoints.item"):
            for station_set in cp.get("stationSets", []):
                for s in station_set:
                    if s.get("codeList") == "UIC":
                        uic = str(s["code"])
                        if uic not in seen:
                            seen.add(uic)
                            stations.append({
                                "cp_id": cp["id"],
                                "uic": uic,
                                "name": uic_to_name.get(uic, uic),
                                "country": s.get("country", ""),
                            })
        stations.sort(key=lambda x: x["name"].lower())

        job["percent"] = 65

        carriers: list[dict] = []
        seen_codes: set[str] = set()
        for cc in ijson.items(_s(), "fareDelivery.fareStructure.carrierConstraints.item"):
            for code in cc.get("includedCarrier", []):
                if code not in seen_codes:
                    seen_codes.add(code)
                    carriers.append({
                        "code": code,
                        "name": RICS_CARRIER_NAMES.get(code, code),
                        "constraint_id": cc["id"],
                    })

        job["percent"] = 75

        seen_refs: dict[str, dict] = {}
        for pc in ijson.items(_s(), "fareDelivery.fareStructure.passengerConstraints.item"):
            ref = pc.get("nameRef", "")
            text_obj = texts_map.get(ref, {})
            name = text_obj.get("textUtf8") or text_obj.get("text") or ref
            if ref not in seen_refs:
                seen_refs[ref] = {"nameRef": ref, "name": name, "ids": []}
            seen_refs[ref]["ids"].append(pc["id"])
        passenger_constraints = list(seen_refs.values())

        job["percent"] = 85

        service_classes = []
        for scd in ijson.items(_s(), "fareDelivery.fareStructure.serviceClassDefinitions.item"):
            text_obj = texts_map.get(scd.get("textRef", ""), {})
            name = text_obj.get("textUtf8") or text_obj.get("text") or scd["id"]
            service_classes.append({"id": scd["id"], "name": name})

        job["percent"] = 95

        job.update({
            "status": "done", "percent": 100,
            "result": {
                "deliveryId": delivery.get("deliveryId", ""),
                "stations": stations,
                "carriers": carriers,
                "passengerConstraints": passenger_constraints,
                "serviceClasses": service_classes,
            }
        })

    except Exception as e:
        job.update({"status": "error", "error": str(e)})


@app.post("/fare-discount/parse")
async def fare_discount_parse(request: Request, osdmFile: UploadFile = File(...)):
    require_login(request)
    file_bytes = await osdmFile.read()
    job_id = str(uuid.uuid4())
    PARSE_JOBS[job_id] = {
        "status": "running", "percent": 0, "phase": "parsing",
        "start_time": time.time(), "file_size": len(file_bytes),
        "result": None, "error": None,
        "owner": request.session.get("user_email"), "created_at": time.time(),
    }
    threading.Thread(target=_run_fare_discount_parse, args=(job_id, file_bytes), daemon=True).start()
    return {"jobId": job_id}


@app.get("/fare-discount/parse/progress/{job_id}")
def fare_discount_parse_progress(job_id: str, request: Request):
    require_login(request)
    job = PARSE_JOBS.get(job_id)
    if not job:
        return {"status": "error", "error": "Jobb ikke funnet"}
    if job.get("owner") != request.session.get("user_email"):
        raise HTTPException(status_code=403, detail="Ikke tilgang")

    if job.get("phase") == "parsing":
        elapsed = time.time() - job.get("start_time", time.time())
        file_mb = job.get("file_size", 0) / (1024 * 1024)
        estimated_s = max(2.0, file_mb / 20)
        pct = min(26, int(elapsed / estimated_s * 26))
        return {"status": "running", "percent": pct, "stage": "reading"}

    if job["status"] in ("done", "error"):
        result = {
            "status": job["status"],
            "percent": job.get("percent", 100),
            "result": job.get("result"),
            "error": job.get("error"),
        }
        PARSE_JOBS.pop(job_id, None)
        return result

    return {"status": "running", "percent": job.get("percent", 0), "stage": "validating"}


# ---------------------------------------------------------------------------
# Hjelpefunksjoner for fare-discount/apply
# ---------------------------------------------------------------------------

def _round_up_020(eur: float) -> float:
    """Rund opp til nærmeste 0.20 EUR."""
    return math.ceil(eur / 0.20) * 0.20


def _id_base(data: dict) -> str:
    """Utled ID-base (f.eks. '1076_7.0_') fra fareProvider og deliveryId i filen."""
    delivery = data.get("fareDelivery", {}).get("delivery", {})
    provider = delivery.get("fareProvider", "")
    did = delivery.get("deliveryId", "")
    if provider and did:
        return f"{provider}_{did}_"
    # Fallback: trekk ut fra eksisterende IDer i filen
    fs = data.get("fareDelivery", {}).get("fareStructure", {})
    for sample in [
        next((c["id"] for c in fs.get("carrierConstraints", [])), None),
        next((t["id"] for t in fs.get("texts", [])), None),
        next((p["id"] for p in fs.get("prices", [])), None),
    ]:
        if sample and "_" in sample:
            parts = sample.split("_")
            if len(parts) >= 2:
                return "_".join(parts[:2]) + "_"
    return f"disc_{did}_" if did else "disc_"


def _id_base_from_delivery(delivery: dict, sample_ids: list[str]) -> str:
    """Utled ID-base fra delivery dict og en liste med eksisterende IDer som fallback."""
    provider = delivery.get("fareProvider", "")
    did = delivery.get("deliveryId", "")
    if provider and did:
        return f"{provider}_{did}_"
    for sample in sample_ids:
        if sample and "_" in sample:
            parts = sample.split("_")
            if len(parts) >= 2:
                return "_".join(parts[:2]) + "_"
    return f"disc_{did}_" if did else "disc_"


def _next_id_num(existing_ids: list[str], prefix: str) -> int:
    """Finn neste ledige nummer for IDer med gitt prefix."""
    import re
    pattern = re.compile(re.escape(prefix) + r"(\d+)$")
    nums = [int(m.group(1)) for id_ in existing_ids if (m := pattern.search(id_))]
    return max(nums) + 1 if nums else 1


def _new_fare_id() -> str:
    """Generer en unik fare-ID i samme format som eksisterende IDer."""
    return "_" + base64.urlsafe_b64encode(uuid.uuid4().bytes).rstrip(b"=").decode()


@app.post("/fare-discount/apply")
async def fare_discount_apply(
    request: Request,
    osdmFile: UploadFile = File(...),
    stationPairsJson: str = Form(...),
    discountName: str = Form(...),
    carrierCodes: list[str] = Form(default=[]),
    discountPct: float = Form(...),
    passengerRefs: list[str] = Form(...),
    serviceClassIds: list[str] = Form(...),
):
    require_login(request)

    try:
        content = await osdmFile.read()
    except Exception:
        raise HTTPException(status_code=400, detail="Filen kunne ikke leses")

    try:
        station_pairs = json.loads(stationPairsJson)
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig stationPairsJson")

    # Quick structure check
    has_fs = False
    try:
        for prefix, event, _ in ijson.parse(io.BytesIO(content)):
            if prefix == "fareDelivery.fareStructure" and event == "start_map":
                has_fs = True
                break
    except Exception:
        raise HTTPException(status_code=400, detail="Filen er ikke gyldig JSON")
    if not has_fs:
        raise HTTPException(status_code=400, detail="Filen mangler fareDelivery.fareStructure")

    # Load small sections via ijson
    delivery: dict = {}
    for item in ijson.items(io.BytesIO(content), "fareDelivery.delivery"):
        delivery = item
        break
    delivery_id = delivery.get("deliveryId", "")

    regional_constraints: list[dict] = list(ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.regionalConstraints.item"))
    passenger_constraints: list[dict] = list(ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.passengerConstraints.item"))
    prices_list: list[dict] = list(ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.prices.item"))
    texts_list: list[dict] = list(ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.texts.item"))
    carrier_constraints_list: list[dict] = list(ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.carrierConstraints.item"))

    # Derive id_base from delivery + sample IDs
    sample_ids = (
        [carrier_constraints_list[0]["id"]] if carrier_constraints_list else
        [texts_list[0]["id"]] if texts_list else
        [prices_list[0]["id"]] if prices_list else
        []
    )
    id_base = _id_base_from_delivery(delivery, sample_ids)

    # Finn relevante RC-er
    if not station_pairs:
        matching_rc_ids = {rc["id"] for rc in regional_constraints}
        if not matching_rc_ids:
            raise HTTPException(status_code=400, detail="OSDM-filen inneholder ingen regionalConstraints")
    else:
        matching_rc_ids: set[str] = set()
        for pair in station_pairs:
            pair_cps = {pair["fromCpId"], pair["toCpId"]}
            for rc in regional_constraints:
                if {rc.get("entryConnectionPointId"), rc.get("exitConnectionPointId")} == pair_cps:
                    matching_rc_ids.add(rc["id"])
        if not matching_rc_ids:
            raise HTTPException(status_code=400, detail="Ingen regionalConstraints funnet for valgte stasjonspar")

    # nameRef → liste av passengerConstraint-IDer
    nameref_to_pc_ids: dict[str, list[str]] = {}
    for pc in passenger_constraints:
        ref = pc.get("nameRef", "")
        if ref in passengerRefs:
            nameref_to_pc_ids.setdefault(ref, []).append(pc["id"])
    selected_pc_ids = {pid for ids in nameref_to_pc_ids.values() for pid in ids}

    # Prisoppslag: priceId → første price-element
    price_lookup: dict[str, dict] = {
        p["id"]: p["price"][0] for p in prices_list if p.get("price")
    }

    multiplier = 1 - discountPct / 100

    # --- Ny carrierConstraint (kun hvis transportører er valgt) ---
    new_cc_id: str | None = None
    new_carrier_constraint: dict | None = None
    if carrierCodes:
        cc_prefix = f"{id_base}C__"
        existing_cc_ids = [c["id"] for c in carrier_constraints_list]
        new_cc_id = f"{cc_prefix}{_next_id_num(existing_cc_ids, cc_prefix)}"
        new_carrier_constraint = {"id": new_cc_id, "includedCarrier": list(carrierCodes)}

    # --- Ny tekst ---
    text_prefix = f"{id_base}P__"
    existing_text_ids = [t["id"] for t in texts_list]
    new_text_id = f"{text_prefix}{_next_id_num(existing_text_ids, text_prefix)}"
    new_text = {
        "id": new_text_id,
        "textUtf8": discountName,
        "text": discountName,
        "shortTextUtf8": discountName,
        "shortText": discountName,
        "translations": [],
    }

    # --- Nye priser (dedupliser på beløp) ---
    price_prefix = f"{id_base}I__"
    existing_price_ids = [p["id"] for p in prices_list]
    next_price_num = _next_id_num(existing_price_ids, price_prefix)
    new_amount_to_price_id: dict[int, str] = {}
    new_prices: list[dict] = []

    def get_or_create_price_id(orig: dict) -> str:
        nonlocal next_price_num
        scale = orig.get("scale", 2)
        eur = orig["amount"] / (10 ** scale)
        discounted_int = int(round(_round_up_020(eur * multiplier) * (10 ** scale)))
        if discounted_int in new_amount_to_price_id:
            return new_amount_to_price_id[discounted_int]
        new_id = f"{price_prefix}{next_price_num}"
        next_price_num += 1
        new_prices.append({
            "id": new_id,
            "price": [{"currency": orig.get("currency", "EUR"), "amount": discounted_int,
                        "scale": scale, "vatDetails": []}],
        })
        new_amount_to_price_id[discounted_int] = new_id
        return new_id

    # --- Nye farer (stream fares, never load full array) ---
    new_fares: list[dict] = []
    seen_combos: set[tuple] = set()

    for fare in ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.fares.item"):
        rc_ref = fare.get("regionalConstraintRef")
        pc_ref = fare.get("passengerConstraintRef")
        sc_ref = fare.get("serviceClassRef")

        if rc_ref not in matching_rc_ids:
            continue
        if pc_ref not in selected_pc_ids:
            continue
        if sc_ref not in serviceClassIds:
            continue

        key = (rc_ref, pc_ref, sc_ref)
        if key in seen_combos:
            continue
        seen_combos.add(key)

        orig_price = price_lookup.get(fare.get("priceRef", ""))
        if not orig_price:
            continue

        new_fare: dict = {
            "id": _new_fare_id(),
            "bundleRef": fare.get("bundleRef", ""),
            "fareType": fare.get("fareType", "ADMISSION"),
            "nameRef": new_text_id,
            "priceRef": get_or_create_price_id(orig_price),
            "regionalConstraintRef": rc_ref,
            "regulatoryConditions": fare.get("regulatoryConditions", ["CIV"]),
            "serviceClassRef": sc_ref,
            "passengerConstraintRef": pc_ref,
            "involvedTCOs": list(carrierCodes),
        }
        if new_cc_id:
            new_fare["carrierConstraintRef"] = new_cc_id
        new_fares.append(new_fare)

    if not new_fares:
        raise HTTPException(
            status_code=400,
            detail="Ingen eksisterende farer funnet for valgt kombinasjon av stasjoner, passasjerer og serviceklasse",
        )

    # Append new items to arrays using byte-level operations
    result_bytes = content
    result_bytes = _append_to_json_array(result_bytes, "fares", new_fares)
    result_bytes = _append_to_json_array(result_bytes, "prices", new_prices)
    result_bytes = _append_to_json_array(result_bytes, "texts", [new_text])
    if new_carrier_constraint:
        result_bytes = _append_to_json_array(result_bytes, "carrierConstraints", [new_carrier_constraint])

    log_event(request.session.get("user_email"), "discount_applied", detail={
        "deliveryId": delivery_id,
        "stationPairs": station_pairs,
        "carrierCodes": list(carrierCodes),
        "discountPct": discountPct,
        "fareName": discountName,
        "fareCount": len(new_fares),
        "priceCount": len(new_prices),
    })

    filename = f"fareDelivery_{id_base.rstrip('_')}_discount.json"
    return Response(
        content=result_bytes,
        media_type="application/json",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Fare-Count": str(len(new_fares)),
            "X-Price-Count": str(len(new_prices)),
        },
    )


def _ijson_default(obj):
    if isinstance(obj, decimal.Decimal):
        return int(obj) if obj == obj.to_integral_value() else float(obj)
    raise TypeError(f"Object of type {obj.__class__.__name__} is not JSON serializable")


@app.post("/ui/fix-osdm")
async def fix_osdm(request: Request, osdmFile: UploadFile = File(...)):
    require_login(request)

    try:
        content = await osdmFile.read()
    except Exception:
        raise HTTPException(status_code=400, detail="Filen kunne ikke leses")

    # Quick structure check
    has_fs = False
    try:
        for prefix, event, _ in ijson.parse(io.BytesIO(content)):
            if prefix == "fareDelivery.fareStructure" and event == "start_map":
                has_fs = True
                break
    except Exception:
        raise HTTPException(status_code=400, detail="Filen er ikke gyldig JSON")
    if not has_fs:
        raise HTTPException(status_code=400, detail="Filen mangler fareDelivery.fareStructure")

    stats: dict[str, int] = {}

    # Bygg sett med definerte IDer via ijson
    price_ids  = {p["id"] for p in ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.prices.item")}
    pc_ids     = {p["id"] for p in ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.passengerConstraints.item")}
    cp_ids     = {cp["id"] for cp in ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.connectionPoints.item")}
    cc_ids     = {c["id"] for c in ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.carrierConstraints.item")}
    bundle_ids = {b["id"] for b in ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.fareConstraintBundles.item")}
    text_ids   = {t["id"] for t in ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.texts.item")}

    # Load regionalConstraints list (small)
    rc_list_orig: list[dict] = list(ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.regionalConstraints.item"))

    # 1. Fjern RC-er med ugyldig CP-referanse
    rc_list_good = [
        rc for rc in rc_list_orig
        if rc.get("entryConnectionPointId") in cp_ids
        and rc.get("exitConnectionPointId") in cp_ids
    ]
    rc_ids = {r["id"] for r in rc_list_good}
    stats["removed_bad_rcs"] = len(rc_list_orig) - len(rc_list_good)

    # 2. Stream fares: find bad fares AND collect used IDs simultaneously
    fares_count_orig = 0
    bad_fare_ids: set[str] = set()
    used_price_ids: set[str] = set()
    used_pc_ids: set[str] = set()
    used_rc_ids: set[str] = set()

    for fare in ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.fares.item"):
        fares_count_orig += 1
        is_bad = (
            (fare.get("priceRef") and fare["priceRef"] not in price_ids)
            or (fare.get("passengerConstraintRef") and fare["passengerConstraintRef"] not in pc_ids)
            or (fare.get("regionalConstraintRef") and fare["regionalConstraintRef"] not in rc_ids)
            or (fare.get("carrierConstraintRef") and fare["carrierConstraintRef"] not in cc_ids)
            or (fare.get("bundleRef") and fare["bundleRef"] not in bundle_ids)
            or (fare.get("nameRef") and fare["nameRef"] not in text_ids)
        )
        if is_bad:
            bad_fare_ids.add(fare.get("id", ""))
        else:
            if fare.get("priceRef"):
                used_price_ids.add(fare["priceRef"])
            if fare.get("passengerConstraintRef"):
                used_pc_ids.add(fare["passengerConstraintRef"])
            if fare.get("regionalConstraintRef"):
                used_rc_ids.add(fare["regionalConstraintRef"])

    bad_fare_count = len(bad_fare_ids)
    stats["removed_bad_fares"] = bad_fare_count

    # Load small arrays for filtering
    prices_orig: list[dict] = list(ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.prices.item"))
    pc_orig: list[dict] = list(ijson.items(io.BytesIO(content), "fareDelivery.fareStructure.passengerConstraints.item"))

    prices_filtered = [p for p in prices_orig if p["id"] in used_price_ids]
    pc_filtered = [p for p in pc_orig if p["id"] in used_pc_ids]
    rc_filtered = [r for r in rc_list_good if r["id"] in used_rc_ids]

    stats["removed_unused_prices"] = len(prices_orig) - len(prices_filtered)
    stats["removed_unused_pcs"] = len(pc_orig) - len(pc_filtered)
    stats["removed_unused_rcs"] = len(rc_list_good) - len(rc_filtered)

    # Build output bytes
    if bad_fare_count > 0:
        # Rebuild fares section by streaming and filtering
        fares_val_start, fares_val_end = _find_section_span(content, "fares")
        out = bytearray()
        out += content[:fares_val_start]
        out += b"["
        first = True
        for fare in ijson.items(io.BytesIO(content[fares_val_start:fares_val_end]), "item"):
            fare_id = fare.get("id", "")
            if fare_id in bad_fare_ids:
                continue
            if not first:
                out += b","
            out += json.dumps(fare, ensure_ascii=False, default=_ijson_default).encode("utf-8")
            first = False
        out += b"]"
        out += content[fares_val_end:]
        result_bytes = bytes(out)
    else:
        result_bytes = content

    # Apply replacements for the small arrays
    try:
        prices_span = _find_section_span(result_bytes, "prices")
        pc_span     = _find_section_span(result_bytes, "passengerConstraints")
        rc_span     = _find_section_span(result_bytes, "regionalConstraints")
    except KeyError as e:
        raise HTTPException(status_code=500, detail=f"Intern feil: mangler nøkkel {e}")

    result_bytes = _apply_byte_replacements(result_bytes, [
        (prices_span[0], prices_span[1], json.dumps(prices_filtered, ensure_ascii=False, default=_ijson_default).encode("utf-8")),
        (pc_span[0],     pc_span[1],     json.dumps(pc_filtered,     ensure_ascii=False, default=_ijson_default).encode("utf-8")),
        (rc_span[0],     rc_span[1],     json.dumps(rc_filtered,     ensure_ascii=False, default=_ijson_default).encode("utf-8")),
    ])

    orig_name = _safe_filename(osdmFile.filename or "osdm.json")
    base = orig_name.rsplit(".", 1)[0] if "." in orig_name else orig_name
    out_name = f"{base}_fixed.json"

    user_email = request.session.get("user_email")
    FIX_OSDM_STORE[user_email] = {"filename": out_name, "content": result_bytes, "created_at": time.time()}

    log_event(user_email, "osdm_analyzed", detail=stats)

    from fastapi.responses import JSONResponse as _JSONResponse
    return _JSONResponse({"stats": stats, "filename": out_name})


@app.get("/ui/fix-osdm/download")
async def fix_osdm_download(request: Request):
    user_email = request.session.get("user_email")
    if not user_email:
        raise HTTPException(status_code=401, detail="Ikke innlogget")
    entry = FIX_OSDM_STORE.get(user_email)
    if not entry:
        raise HTTPException(status_code=404, detail="Ingen fikset fil funnet — last opp og analyser først")
    log_event(user_email, "osdm_fixed", detail={"filename": entry["filename"]})
    return Response(
        content=entry["content"],
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{entry["filename"]}"'},
    )


@app.get("/osdmtoexcel", response_class=HTMLResponse)
@app.head("/osdmtoexcel")
def osdmtoexcel_page(request: Request):
    if "user_email" not in request.session:
        return HTMLResponse(Path("frontend/login.html").read_text(encoding="utf-8"))
    is_admin = bool(request.session.get("is_admin"))
    html = Path("frontend/osdmtoexcel.html").read_text(encoding="utf-8")
    html = html.replace(
        "</head>",
        f"<script>window.IS_ADMIN = {str(is_admin).lower()};</script></head>"
    )
    return HTMLResponse(html)

@app.get("/fix-osdm", response_class=HTMLResponse)
@app.head("/fix-osdm")
def fix_osdm_page(request: Request):
    if "user_email" not in request.session:
        return HTMLResponse(Path("frontend/login.html").read_text(encoding="utf-8"))
    is_admin = bool(request.session.get("is_admin"))
    html = Path("frontend/fix-osdm.html").read_text(encoding="utf-8")
    html = html.replace(
        "</head>",
        f"<script>window.IS_ADMIN = {str(is_admin).lower()};</script></head>"
    )
    return HTMLResponse(html)
